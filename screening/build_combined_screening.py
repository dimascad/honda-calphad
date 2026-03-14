"""
Build unified screening workbook combining binary oxide and ternary reaction data.

Tabs:
  1. Formation Energies      — raw dGf from TC-Python (from oxide_gibbs_energies.csv)
  2. Binary Screening        — Cu2O vs MOx Ellingham comparison + verdict
  3. Ternary Reaction Balance — atom balance for Cu + MOx -> CuMOy (always available)
  4. Ternary Reaction dG      — dG_rxn at key temperatures (needs ternary CSV)
  5. Combined Screening       — one row per oxide, binary + ternary verdicts side by side
  6. Normalization Explained  — how GM values are converted to per-mol-O2
  7. Pricing & SDS            — cost, CAS, safety data

Run locally. Reads from:
  - data/tcpython/raw/oxide_gibbs_energies.csv        (binary oxide data)
  - data/tcpython/raw/ternary_reaction_energies.csv   (ternary data, optional)

Output:
  - screening/Cu_Removal_Screening.xlsx
"""

import csv
from pathlib import Path

import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SCRIPT_DIR = Path(__file__).parent
DATA_DIR = SCRIPT_DIR.parent / "data" / "tcpython" / "raw"
BINARY_CSV = DATA_DIR / "oxide_gibbs_energies.csv"
TERNARY_CSV = DATA_DIR / "ternary_reaction_energies.csv"
OUTPUT_FILE = SCRIPT_DIR / "Cu_Removal_Screening.xlsx"

# Key temperatures
TEMPS_K = [1000, 1500, 1800]  # 727, 1227, 1527 C

# ============================================================================
# Style definitions (IEEE standard from CLAUDE.md)
# ============================================================================
header_fill = PatternFill(start_color="595959", end_color="595959", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
alt_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
tbd_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
green_font = Font(color="006100", bold=True, size=11)
red_font = Font(color="9C0006", bold=True, size=11)
bold_font = Font(bold=True, size=11)
normal_font = Font(size=11)
small_font = Font(size=10, color="666666")
small_italic = Font(size=10, color="666666", italic=True)
tbd_font = Font(italic=True, color="996600", size=11)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)

thick = Side(style="medium")
thin = Side(style="thin")


def apply_borders(ws, row_start, row_end, col_start, col_end,
                  group_breaks=None):
    """Apply IEEE-style borders to a table region."""
    if group_breaks is None:
        group_breaks = set()
    for r in range(row_start, row_end + 1):
        for c in range(col_start, col_end + 1):
            cell = ws.cell(row=r, column=c)
            l = thick if c == col_start else thin
            ri = thick if c == col_end else thin
            t = thick if r == row_start else thin
            b = thick if r == row_end else thin
            if r == row_start:
                b = thick
            if r in group_breaks:
                b = thick
            if r - 1 in group_breaks and r != row_start:
                t = thick
            cell.border = Border(left=l, right=ri, top=t, bottom=b)


def add_notes(ws, start_row, notes):
    """Add a notes section below a table."""
    ws.cell(row=start_row, column=1, value="Notes:").font = bold_font
    for i, note in enumerate(notes, 1):
        ws.cell(row=start_row + i, column=1, value=note).font = small_italic


# ============================================================================
# Oxide physical properties (hardcoded reference data)
# ============================================================================
OXIDE_PROPS = {
    "MgO":    {"density": 3.58, "Tm_C": 2852, "toxicity": "Low", "cost_ind": 0.0003},
    "Al2O3":  {"density": 3.95, "Tm_C": 2072, "toxicity": "Low", "cost_ind": 0.0005},
    "CaO":    {"density": 3.34, "Tm_C": 2613, "toxicity": "Low", "cost_ind": 0.0002},
    "ZrO2":   {"density": 5.68, "Tm_C": 2715, "toxicity": "Low", "cost_ind": 0.035},
    "Cr2O3":  {"density": 5.22, "Tm_C": 2435, "toxicity": "Moderate (Cr VI)", "cost_ind": 0.005},
    "TiO2":   {"density": 4.23, "Tm_C": 1843, "toxicity": "Low", "cost_ind": 0.003},
    "SiO2":   {"density": 2.65, "Tm_C": 1713, "toxicity": "Low (silicosis)", "cost_ind": 5e-05},
    "MnO":    {"density": 5.43, "Tm_C": 1842, "toxicity": "Low", "cost_ind": 0.005},
    "V2O5":   {"density": 3.36, "Tm_C": 690,  "toxicity": "HIGH - Carc. 1B", "cost_ind": 0.011},
    "FeO":    {"density": 5.74, "Tm_C": 1377, "toxicity": "Low", "cost_ind": 0.0009},
    "CoO":    {"density": 6.44, "Tm_C": 1830, "toxicity": "HIGH - Carc. 1B", "cost_ind": 0.018},
    "NiO":    {"density": 6.67, "Tm_C": 1955, "toxicity": "HIGH - IARC Grp 1", "cost_ind": 0.016},
    "PbO":    {"density": 9.53, "Tm_C": 888,  "toxicity": "HIGH - neurotoxin", "cost_ind": 0.004},
    "B2O3":   {"density": 2.55, "Tm_C": 450,  "toxicity": "Moderate - Repr.1B", "cost_ind": 0.003},
    "La2O3":  {"density": 6.51, "Tm_C": 2315, "toxicity": "Low-Moderate", "cost_ind": 0.01},
    "CeO2":   {"density": 7.22, "Tm_C": 2400, "toxicity": "Low", "cost_ind": 0.006},
    "CuO":    {"density": 6.31, "Tm_C": 1326, "toxicity": "Moderate - aquatic", "cost_ind": 0.008},
    "Cu2O":   {"density": 6.0,  "Tm_C": 1235, "toxicity": "Reference", "cost_ind": 0.008},
}

# Column names in the binary oxide CSV for dG per mol O2
OXIDE_DG_COLS = {
    "Cu2O":  "dG_Cu2O_per_O2",
    "CuO":   "dG_CuO_per_O2",
    "Al2O3": "dG_Al2O3_per_O2",
    "MgO":   "dG_MgO_per_O2",
    "SiO2":  "dG_SiO2_per_O2",
    "TiO2":  "dG_TiO2_per_O2",
    "FeO":   "dG_FeO_per_O2",
    "CaO":   "dG_CaO_per_O2",
    "ZrO2":  "dG_ZrO2_per_O2",
    "Cr2O3": "dG_Cr2O3_per_O2",
    "MnO":   "dG_MnO_per_O2",
    "NiO":   "dG_NiO_per_O2",
    "CoO":   "dG_CoO_per_O2",
    "PbO":   "dG_PbO_per_O2",
    "B2O3":  "dG_B2O3_per_O2",
    "V2O5":  "dG_V2O5_per_O2",
    "La2O3": "dG_La2O3_per_O2",
    "CeO2":  "dG_CeO2_per_O2",
}

# Screening order (most stable first, Cu2O/CuO last)
SCREENING_ORDER = [
    "CeO2", "CaO", "La2O3", "MgO", "ZrO2", "Al2O3",
    "TiO2", "SiO2", "MnO", "Cr2O3", "V2O5", "FeO",
    "CoO", "NiO", "PbO", "B2O3", "CuO",
]

# ============================================================================
# Ternary reaction definitions (same as extract_ternary_reactions.py)
# ============================================================================
TERNARY_REACTIONS = [
    # (product, name, reaction_str, oxide, n_Cu, n_oxide, ox_formula, ox_atoms, n_O2, prod_atoms, category)
    ("CuAl2O4",  "Copper aluminate spinel",       "Cu + Al2O3 + 0.5 O2 -> CuAl2O4",     "Al2O3", 1, 1,   "Al2O3", 5, 0.5,  7, "Spinel"),
    ("CuAlO2",   "Delafossite",                   "Cu + 0.5 Al2O3 + 0.25 O2 -> CuAlO2", "Al2O3", 1, 0.5, "Al2O3", 5, 0.25, 4, "Delafossite"),
    ("CuCr2O4",  "Copper chromite spinel",         "Cu + Cr2O3 + 0.5 O2 -> CuCr2O4",    "Cr2O3", 1, 1,   "Cr2O3", 5, 0.5,  7, "Spinel"),
    ("CuMn2O4",  "Copper manganite spinel",        "Cu + 2MnO + O2 -> CuMn2O4",          "MnO",   1, 2,   "MnO",   2, 1.0,  7, "Spinel"),
    ("CuFe2O4",  "Copper ferrite spinel",          "Cu + 2FeO + O2 -> CuFe2O4",          "FeO",   1, 2,   "FeO",   2, 1.0,  7, "Spinel"),
    ("CuCo2O4",  "Copper cobaltite spinel",        "Cu + 2CoO + O2 -> CuCo2O4",          "CoO",   1, 2,   "CoO",   2, 1.0,  7, "Spinel"),
    ("CuV2O6",   "Copper vanadate",                "Cu + V2O5 + 0.5 O2 -> CuV2O6",      "V2O5",  1, 1,   "V2O5",  7, 0.5,  9, "Vanadate"),
    ("Cu3V2O8",  "Copper orthovanadate",           "3Cu + V2O5 + 1.5 O2 -> Cu3V2O8",    "V2O5",  3, 1,   "V2O5",  7, 1.5, 13, "Vanadate"),
    ("CuTiO3",   "Copper titanate",                "Cu + TiO2 + 0.5 O2 -> CuTiO3",      "TiO2",  1, 1,   "TiO2",  3, 0.5,  5, "Titanate"),
    ("CuSiO3",   "Copper metasilicate",            "Cu + SiO2 + 0.5 O2 -> CuSiO3",      "SiO2",  1, 1,   "SiO2",  3, 0.5,  5, "Silicate"),
    ("Cu2SiO4",  "Copper orthosilicate",           "2Cu + SiO2 + O2 -> Cu2SiO4",         "SiO2",  2, 1,   "SiO2",  3, 1.0,  7, "Silicate"),
    ("CuMgO2",   "Copper magnesioxide (hyp.)",     "Cu + MgO + 0.5 O2 -> CuMgO2",       "MgO",   1, 1,   "MgO",   2, 0.5,  4, "Hypothetical"),
    ("CuCaO2",   "Copper calcioxide (hyp.)",       "Cu + CaO + 0.5 O2 -> CuCaO2",       "CaO",   1, 1,   "CaO",   2, 0.5,  4, "Hypothetical"),
    ("CuZrO3",   "Copper zirconate (hyp.)",        "Cu + ZrO2 + 0.5 O2 -> CuZrO3",      "ZrO2",  1, 1,   "ZrO2",  3, 0.5,  5, "Hypothetical"),
    ("CuNiO2",   "Copper nickelate",               "Cu + NiO + 0.5 O2 -> CuNiO2",       "NiO",   1, 1,   "NiO",   2, 0.5,  4, "Compound"),
    ("CuLaO2",   "Copper lanthanum oxide",         "Cu + 0.5 La2O3 + 0.25 O2 -> CuLaO2","La2O3", 1, 0.5, "La2O3", 5, 0.25, 4, "Compound"),
    ("CuCeO3",   "Copper cerate (hyp.)",           "Cu + CeO2 + 0.5 O2 -> CuCeO3",      "CeO2",  1, 1,   "CeO2",  3, 0.5,  5, "Hypothetical"),
    ("CuB2O4",   "Copper borate",                  "Cu + B2O3 + 0.5 O2 -> CuB2O4",      "B2O3",  1, 1,   "B2O3",  5, 0.5,  7, "Borate"),
]

# Phase hints: what TC phase names correspond to each product
PHASE_HINTS = {
    "CuAl2O4": ["SPINEL"], "CuAlO2": ["DELAFOSSITE"],
    "CuCr2O4": ["SPINEL"], "CuMn2O4": ["SPINEL"],
    "CuFe2O4": ["SPINEL"], "CuCo2O4": ["SPINEL"],
    "CuV2O6": ["VANADATE", "CUV2O6"], "Cu3V2O8": ["ORTHOVANADATE", "CU3V2O8"],
    "CuTiO3": ["ILMENITE", "PEROVSKITE"], "CuSiO3": ["PYROXENE", "WOLLASTONITE"],
    "Cu2SiO4": ["OLIVINE"], "CuMgO2": ["DELAFOSSITE"],
    "CuCaO2": ["DELAFOSSITE"], "CuZrO3": ["PEROVSKITE"],
    "CuNiO2": ["DELAFOSSITE"], "CuLaO2": ["DELAFOSSITE"],
    "CuCeO3": ["PEROVSKITE"], "CuB2O4": ["BORATE"],
}

# Pricing & SDS data
SDS_DATA = [
    ("CeO2",  "1306-38-3",  0.90,  0.006,  "211575"),
    ("CaO",   "1305-78-8",  0.16,  0.0002, "248568"),
    ("La2O3", "1312-81-8",  1.25,  0.01,   "L4000"),
    ("MgO",   "1309-48-4",  0.13,  0.0003, "529699"),
    ("ZrO2",  "1314-23-4",  0.67,  0.035,  "230693"),
    ("Al2O3", "1344-28-1",  0.09,  0.0005, "11028"),
    ("TiO2",  "13463-67-7", 0.29,  0.003,  "14021"),
    ("SiO2",  "7631-86-9",  0.10,  5e-05,  "637238"),
    ("MnO",   "1344-43-0",  0.55,  0.005,  "377201"),
    ("Cr2O3", "1308-38-9",  2.38,  0.005,  "393703"),
    ("V2O5",  "1314-62-1",  1.00,  0.011,  "221899"),
    ("FeO",   "1345-25-1",  0.75,  0.0009, "400866"),
    ("CoO",   "1307-96-6",  1.00,  0.018,  "529443"),
    ("NiO",   "1313-99-1",  0.70,  0.016,  "203882"),
    ("PbO",   "1317-36-8",  0.26,  0.004,  "211907"),
    ("Cu2O",  "1317-39-1",  0.34,  0.008,  "208825"),
    ("CuO",   "1317-38-0",  0.38,  0.008,  "208841"),
    ("B2O3",  "1303-86-2",  1.09,  0.003,  "339075"),
]

# Normalization stoichiometry
NORM_TABLE = [
    ("Cu2O",  3, "4Cu + O2 -> 2Cu2O",              2),
    ("Al2O3", 5, "4/3 Al + O2 -> 2/3 Al2O3",       0.667),
    ("MgO",   2, "2Mg + O2 -> 2MgO",               2),
    ("SiO2",  3, "Si + O2 -> SiO2",                 1),
    ("TiO2",  3, "Ti + O2 -> TiO2",                 1),
    ("FeO",   2, "2Fe + O2 -> 2FeO",                2),
    ("CaO",   2, "2Ca + O2 -> 2CaO",                2),
    ("ZrO2",  3, "Zr + O2 -> ZrO2",                 1),
    ("Cr2O3", 5, "4/3 Cr + O2 -> 2/3 Cr2O3",       0.667),
    ("MnO",   2, "2Mn + O2 -> 2MnO",                2),
    ("V2O5",  7, "4/5 V + O2 -> 2/5 V2O5",          0.4),
    ("B2O3",  5, "4/3 B + O2 -> 2/3 B2O3",          0.667),
    ("La2O3", 5, "4/3 La + O2 -> 2/3 La2O3",        0.667),
    ("CeO2",  3, "Ce + O2 -> CeO2",                 1),
    ("NiO",   2, "2Ni + O2 -> 2NiO",                2),
    ("CoO",   2, "2Co + O2 -> 2CoO",                2),
    ("PbO",   2, "2Pb + O2 -> 2PbO",                2),
]


# ============================================================================
# Data loading
# ============================================================================

def find_ternary_phase_info(product_rows, product):
    """Scan all temperatures to find the named ternary phase and its stability range."""
    hints = PHASE_HINTS.get(product, [])
    if not hints:
        return "", ""

    max_T = None
    found_phase = None
    for r in product_rows:
        phases = r.get("stable_phases", "")
        for hint in hints:
            if hint.upper() in phases.upper():
                T = float(r["T_K"])
                if max_T is None or T > max_T:
                    max_T = T
                    for p in phases.split(";"):
                        if hint.upper() in p.strip().upper():
                            found_phase = p.strip()

    if found_phase and max_T:
        return found_phase, f"Stable up to {int(max_T)}K ({int(max_T)-273}C)"
    return "", ""


def load_binary_data():
    """Load binary oxide Gibbs energy data from CSV."""
    if not BINARY_CSV.exists():
        print(f"WARNING: {BINARY_CSV} not found. Binary tabs will use placeholders.")
        return None
    import pandas as pd
    df = pd.read_csv(BINARY_CSV)
    print(f"Loaded binary data: {len(df)} temperatures, {len(df.columns)} columns")
    return df


def load_ternary_data():
    """Load ternary reaction data from CSV."""
    if not TERNARY_CSV.exists():
        print(f"INFO: {TERNARY_CSV} not found. Ternary dG tab will show TBD.")
        return []
    with open(TERNARY_CSV) as f:
        rows = list(csv.DictReader(f))
    print(f"Loaded ternary data: {len(rows)} rows")
    return rows


# ============================================================================
# Tab 1: Formation Energies
# ============================================================================

def build_tab1_formation_energies(wb, df):
    """Raw dGf per mol O2 at each temperature for all oxides."""
    ws = wb.active
    ws.title = "Formation Energies"

    # Oxide column order (most stable first)
    oxide_cols = [
        "CeO2", "CaO", "La2O3", "MgO", "ZrO2", "Al2O3",
        "TiO2", "SiO2", "MnO", "Cr2O3", "V2O5", "FeO",
        "CoO", "NiO", "PbO", "Cu2O", "CuO", "B2O3",
    ]

    headers = ["T (K)"] + oxide_cols
    col_widths = [8] + [14] * len(oxide_cols)

    for c, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        ws.column_dimensions[get_column_letter(c)].width = w

    if df is None:
        ws.cell(row=2, column=1, value="No data — run extract_oxide_gibbs.py on VM").font = tbd_font
        return

    for r_idx, (_, row) in enumerate(df.iterrows()):
        r = r_idx + 2
        ws.cell(row=r, column=1, value=int(row["T_K"])).font = normal_font
        ws.cell(row=r, column=1).alignment = center

        for c_idx, oxide in enumerate(oxide_cols):
            c = c_idx + 2
            col_name = OXIDE_DG_COLS.get(oxide)
            cell = ws.cell(row=r, column=c)

            if col_name and col_name in df.columns:
                val = row.get(col_name)
                if val is not None and not (isinstance(val, float) and np.isnan(val)):
                    cell.value = round(float(val), 1)
                    cell.number_format = '#,##0.0'
                else:
                    cell.value = None
            else:
                cell.value = None

            cell.font = normal_font
            cell.alignment = center
            cell.fill = white_fill if r_idx % 2 == 0 else alt_fill

    last_row = len(df) + 1
    apply_borders(ws, 1, last_row, 1, len(headers))

    add_notes(ws, last_row + 2, [
        "Values are dGf in kJ/mol O2 (Ellingham convention).",
        "Source: TCOX14 database via TC-Python. PbO and CeO2 from SSUB3.",
        "More negative = more thermodynamically stable oxide.",
    ])

    ws.freeze_panes = "B2"
    ws.row_dimensions[1].height = 30


# ============================================================================
# Tab 2: Binary Screening (Ellingham)
# ============================================================================

def build_tab2_binary_screening(wb, df):
    """Cu2O vs MOx comparison — can Cu reduce the oxide?"""
    ws = wb.create_sheet("Binary Screening")

    headers = [
        "Oxide",
        "Density\n(g/cm3)",
        "Tm (C)",
        "Floats on\nSteel?",
        "Solid at\n1527C?",
        "dG_rxn\n@ 727C\n(kJ/mol O2)",
        "dG_rxn\n@ 1227C\n(kJ/mol O2)",
        "dG_rxn\n@ 1527C\n(kJ/mol O2)",
        "Cu Can\nReduce?",
        "Toxicity",
        "Source",
    ]
    col_widths = [10, 10, 10, 10, 10, 16, 16, 16, 10, 24, 14]

    for c, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        ws.column_dimensions[get_column_letter(c)].width = w

    # Compute dG_rxn = dGf(Cu2O) - dGf(MOx) at each temperature
    cu2o_col = "dG_Cu2O_per_O2"

    for r_idx, oxide in enumerate(SCREENING_ORDER):
        r = r_idx + 2
        props = OXIDE_PROPS.get(oxide, {})
        density = props.get("density", "")
        Tm = props.get("Tm_C", "")
        floats = "Yes" if density and density < 7.0 else "No"
        solid = "Yes" if Tm and Tm > 1527 else "LIQUID"
        tox = props.get("toxicity", "")

        # Determine source
        source = "TCOX14"
        if oxide in ("PbO", "CeO2"):
            source = "SSUB3"

        values = [oxide, density, Tm, floats, solid]

        # dG_rxn at each temperature
        dG_vals = {}
        if df is not None:
            ox_col = OXIDE_DG_COLS.get(oxide)
            if ox_col and ox_col in df.columns and cu2o_col in df.columns:
                for T in TEMPS_K:
                    idx = int(np.argmin(np.abs(df["T_K"].values - T)))
                    cu2o_val = df[cu2o_col].values[idx]
                    ox_val = df[ox_col].values[idx]
                    if not np.isnan(cu2o_val) and not np.isnan(ox_val):
                        dG_vals[T] = (cu2o_val - ox_val) / 1000  # J -> kJ
                    else:
                        dG_vals[T] = None
            else:
                for T in TEMPS_K:
                    dG_vals[T] = None
        else:
            for T in TEMPS_K:
                dG_vals[T] = None

        for T in TEMPS_K:
            values.append(dG_vals.get(T))

        # Verdict
        dG_1527 = dG_vals.get(1800)
        if dG_1527 is not None:
            if dG_1527 > 0:
                reducible = "No"
            elif oxide == "CuO":
                reducible = "Trivially"
            else:
                reducible = "Yes"
        else:
            reducible = "TBD"

        values.extend([reducible, tox, source])

        row_fill = white_fill if r_idx % 2 == 0 else alt_fill

        for c, val in enumerate(values, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.fill = row_fill

            # dG columns
            if c in (6, 7, 8):
                if isinstance(val, (int, float)):
                    cell.number_format = '+#,##0.0;-#,##0.0'
                    cell.font = red_font if val > 0 else green_font
                elif val is None:
                    cell.value = "TBD"
                    cell.fill = tbd_fill
                    cell.font = tbd_font
                cell.alignment = center
            elif c == 5 and val == "LIQUID":
                cell.font = Font(bold=True, color="9C0006", size=11)
                cell.alignment = center
            elif c == 9:
                if val == "No":
                    cell.font = bold_font
                elif val == "TBD":
                    cell.fill = tbd_fill
                    cell.font = tbd_font
                else:
                    cell.font = normal_font
                cell.alignment = center
            elif c in (1, 10, 11):
                cell.alignment = left_wrap
                cell.font = normal_font
            else:
                cell.alignment = center
                cell.font = normal_font

    last_row = len(SCREENING_ORDER) + 1
    apply_borders(ws, 1, last_row, 1, len(headers))

    add_notes(ws, last_row + 2, [
        "dG_rxn = dGf(Cu2O) - dGf(MOx) per mol O2. Positive = Cu CANNOT reduce oxide.",
        "Molten steel density ~ 7.0 g/cm3 at 1600C. Oxides lighter than steel float.",
        "Red values = UNFAVORABLE (oxide too stable). Green = Cu can reduce.",
        "CuO: Cu2O -> CuO is further Cu oxidation, not useful for removal.",
        "Conclusion: Cu cannot directly reduce ANY useful ceramic oxide at steelmaking temps.",
    ])

    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 45


# ============================================================================
# Tab 3: Ternary Reaction Balancing
# ============================================================================

def build_tab3_ternary_balancing(wb):
    """Balanced reactions with atom counts — always available."""
    ws = wb.create_sheet("Ternary Reaction Balance")

    headers = [
        "Product",
        "Name",
        "Category",
        "Balanced Reaction",
        "n(Cu)",
        "Oxide",
        "n(oxide)",
        "Oxide\nAtoms",
        "n(O2)",
        "O atoms\nfrom O2",
        "Total\nReactant\nAtoms",
        "Product\nAtoms",
        "Balanced?",
    ]
    col_widths = [14, 28, 14, 40, 8, 10, 8, 8, 8, 10, 10, 10, 10]

    for c, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        ws.column_dimensions[get_column_letter(c)].width = w

    # Track category groups for alternating shading
    categories_seen = []
    for rxn in TERNARY_REACTIONS:
        cat = rxn[10]
        if not categories_seen or categories_seen[-1] != cat:
            categories_seen.append(cat)

    cat_group_idx = {cat: i for i, cat in enumerate(categories_seen)}
    group_break_rows = set()
    prev_cat = None

    for r_idx, rxn in enumerate(TERNARY_REACTIONS):
        r = r_idx + 2
        product, name, reaction, oxide, n_Cu, n_oxide, ox_formula, ox_atoms, n_O2, prod_atoms, category = rxn

        o_from_O2 = n_O2 * 2
        total_reactant = n_Cu + n_oxide * ox_atoms + o_from_O2
        balanced = "YES" if abs(total_reactant - prod_atoms) < 0.01 else "NO"

        if prev_cat is not None and category != prev_cat:
            group_break_rows.add(r - 1)
        prev_cat = category

        group_num = cat_group_idx.get(category, 0)
        row_fill = alt_fill if group_num % 2 == 1 else white_fill

        values = [product, name, category, reaction,
                  n_Cu, ox_formula, n_oxide, ox_atoms, n_O2, o_from_O2,
                  total_reactant, prod_atoms, balanced]

        for c, val in enumerate(values, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = normal_font
            cell.fill = row_fill
            cell.alignment = left_wrap if c in (1, 2, 3, 4, 6) else center

            if c == 13:
                cell.font = green_font if val == "YES" else red_font

    apply_borders(ws, 1, len(TERNARY_REACTIONS) + 1, 1, 13, group_break_rows)

    add_notes(ws, len(TERNARY_REACTIONS) + 3, [
        "All reactions require oxygen. In steelmaking, O2 comes from dissolved oxygen or atmosphere.",
        "n(oxide) = moles of binary oxide consumed per mole of ternary product.",
        "Oxide Atoms = atoms per formula unit (e.g., Al2O3 = 5).",
        "Total Reactant Atoms must equal Product Atoms for balanced reaction.",
        "Hypothetical compounds have no known crystal structure; TC-Python determines stability.",
    ])

    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 40


# ============================================================================
# Tab 4: Ternary Reaction dG
# ============================================================================

def build_tab4_ternary_dG(wb, ternary_data):
    """dG_rxn at key temperatures for each ternary product."""
    ws = wb.create_sheet("Ternary Reaction dG")

    headers = ["Product", "Name", "Oxide", "Reaction"]
    for T in TEMPS_K:
        headers.append(f"dG_rxn\n@ {T-273}C\n(kJ)")
    headers.extend(["Ternary Phase\n@ 1527C", "Stable Phases\n@ 1527C"])
    col_widths = [14, 28, 10, 40, 14, 14, 14, 20, 40]

    for c, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        ws.column_dimensions[get_column_letter(c)].width = w

    # Organize ternary data by product
    by_product = {}
    for row in ternary_data:
        key = row["product"]
        if key not in by_product:
            by_product[key] = []
        by_product[key].append(row)

    group_break_rows = set()
    prev_cat = None

    for r_idx, rxn_def in enumerate(TERNARY_REACTIONS):
        r = r_idx + 2
        product, name, reaction, oxide = rxn_def[0], rxn_def[1], rxn_def[2], rxn_def[3]
        category = rxn_def[10]

        if prev_cat is not None and category != prev_cat:
            group_break_rows.add(r - 1)
        prev_cat = category

        row_fill = white_fill if r_idx % 2 == 0 else alt_fill
        values = [product, name, oxide, reaction]

        prod_rows = by_product.get(product, [])

        # Extract T, dG pairs
        T_arr, dG_arr = [], []
        for pr in prod_rows:
            try:
                T_arr.append(float(pr["T_K"]))
                dG_str = pr.get("dG_rxn_system_kJ", "").strip()
                dG_arr.append(float(dG_str) if dG_str else None)
            except (ValueError, KeyError):
                pass

        # dG at each key temperature
        for T_target in TEMPS_K:
            if T_arr:
                np_T = np.array([t for t, d in zip(T_arr, dG_arr) if d is not None])
                np_dG = np.array([d for d in dG_arr if d is not None])
                if len(np_T) > 0:
                    idx = int(np.argmin(np.abs(np_T - T_target)))
                    if abs(np_T[idx] - T_target) < 30:
                        values.append(round(np_dG[idx], 1))
                    else:
                        values.append("No data")
                else:
                    values.append("No data")
            else:
                values.append("TBD")

        # Phase info: scan ALL temperatures for the named ternary phase
        phase_name, phase_range = find_ternary_phase_info(prod_rows, product)
        if phase_name:
            values.append(f"{phase_name} ({phase_range})")
        elif prod_rows:
            values.append("Dissolved in ionic liquid")
        else:
            values.append("TBD")

        # Stable phases at 1800K
        rows_1800 = [pr for pr in prod_rows
                     if pr.get("T_K") and abs(float(pr["T_K"]) - 1800) < 30]
        if rows_1800:
            values.append(rows_1800[0].get("stable_phases", ""))
        else:
            values.append("TBD" if not ternary_data else "")

        for c, val in enumerate(values, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.fill = row_fill

            if c in (5, 6, 7) and isinstance(val, (int, float)):
                cell.number_format = '+#,##0.0;-#,##0.0'
                if val < -10:
                    cell.font = green_font
                    cell.fill = green_fill
                elif val < 0:
                    cell.font = Font(color="006100", size=11)
                else:
                    cell.font = red_font
            elif c in (5, 6, 7):
                cell.font = tbd_font
            else:
                cell.font = normal_font

            cell.alignment = left_wrap if c in (1, 2, 4, 8, 9) else center

    apply_borders(ws, 1, len(TERNARY_REACTIONS) + 1, 1, len(headers), group_break_rows)

    add_notes(ws, len(TERNARY_REACTIONS) + 3, [
        "dG_rxn = G(products) - G(reactants)",
        "G(products) = GM_system(ternary composition) x atoms_per_formula",
        "G(reactants) = n_Cu x G(Cu metal) + n_oxide x oxide_atoms x G(binary oxide) + n_O2 x G(O2 gas)",
        "Negative dG = FAVORABLE (oxide captures copper into ternary compound)",
        "Positive dG = UNFAVORABLE (ternary compound does not form spontaneously)",
        "Green cells = favorable.  Red = unfavorable.",
    ])

    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 50


# ============================================================================
# Tab 5: Combined Screening Summary
# ============================================================================

def build_tab5_combined_screening(wb, df, ternary_data):
    """One row per oxide with both binary and ternary verdicts."""
    ws = wb.create_sheet("Combined Screening")

    headers = [
        "Oxide",
        "Density\n(g/cm3)",
        "Tm (C)",
        "Floats?",
        "Solid at\n1527C?",
        "Binary dG_rxn\n@ 1527C\n(kJ/mol O2)",
        "Binary\nVerdict",
        "Best Ternary\nProduct",
        "Ternary dG_rxn\n@ 1527C\n(kJ)",
        "Ternary\nVerdict",
        "Overall\nRecommendation",
        "Toxicity",
    ]
    col_widths = [10, 10, 8, 8, 8, 16, 16, 16, 16, 16, 28, 22]

    for c, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        ws.column_dimensions[get_column_letter(c)].width = w

    # Build ternary lookup: oxide -> best product + dG at 1800K
    by_product = {}
    for row in ternary_data:
        key = row["product"]
        if key not in by_product:
            by_product[key] = []
        by_product[key].append(row)

    ternary_best = {}  # oxide -> (product, dG_1800, phase_info)
    for rxn_def in TERNARY_REACTIONS:
        product, oxide = rxn_def[0], rxn_def[3]
        prod_rows = by_product.get(product, [])

        dG_1800 = None
        for pr in prod_rows:
            try:
                T = float(pr["T_K"])
                if abs(T - 1800) < 30:
                    dG_str = pr.get("dG_rxn_system_kJ", "").strip()
                    if dG_str:
                        dG_1800 = float(dG_str)
            except (ValueError, KeyError):
                pass

        # Get named phase info across all temperatures
        phase_name, phase_range = find_ternary_phase_info(prod_rows, product)
        if phase_name:
            phase_info = f"{phase_name} ({phase_range})"
        elif prod_rows:
            phase_info = "Ionic liquid"
        else:
            phase_info = ""

        if oxide not in ternary_best or (
                dG_1800 is not None and
                (ternary_best[oxide][1] is None or dG_1800 < ternary_best[oxide][1])):
            ternary_best[oxide] = (product, dG_1800, phase_info)

    # Binary dG lookup
    cu2o_col = "dG_Cu2O_per_O2"

    for r_idx, oxide in enumerate(SCREENING_ORDER):
        r = r_idx + 2
        props = OXIDE_PROPS.get(oxide, {})
        density = props.get("density", "")
        Tm = props.get("Tm_C", "")
        floats = "Yes" if density and density < 7.0 else "No"
        solid = "Yes" if Tm and Tm > 1527 else "LIQUID"
        tox = props.get("toxicity", "")

        # Binary dG at 1527C
        binary_dG = None
        if df is not None:
            ox_col = OXIDE_DG_COLS.get(oxide)
            if ox_col and ox_col in df.columns and cu2o_col in df.columns:
                idx = int(np.argmin(np.abs(df["T_K"].values - 1800)))
                cu2o_val = df[cu2o_col].values[idx]
                ox_val = df[ox_col].values[idx]
                if not np.isnan(cu2o_val) and not np.isnan(ox_val):
                    binary_dG = (cu2o_val - ox_val) / 1000

        if binary_dG is not None:
            binary_verdict = "CANNOT reduce" if binary_dG > 0 else "CAN reduce"
        else:
            binary_verdict = "TBD"

        # Ternary info
        tern_info = ternary_best.get(oxide)
        if tern_info:
            tern_product, tern_dG, tern_phase = tern_info
            if tern_dG is not None:
                if tern_dG < -10:
                    tern_verdict = "FAVORABLE"
                elif tern_dG < 0:
                    tern_verdict = "Marginal"
                elif tern_dG < 10:
                    tern_verdict = "Borderline"
                else:
                    tern_verdict = "Unfavorable"
            else:
                tern_verdict = "TBD"
        else:
            tern_product = "N/A"
            tern_dG = None
            tern_verdict = "Not modeled"

        # Overall recommendation
        if tern_dG is not None and tern_dG < -10:
            recommendation = "PROMISING — ternary compound forms"
        elif tern_dG is not None and tern_dG < 0:
            recommendation = "Marginal — weak compound formation"
        elif oxide == "CuO":
            recommendation = "Trivial — further Cu oxidation"
        elif tern_verdict == "TBD":
            recommendation = "Awaiting ternary data"
        elif tern_verdict == "Not modeled":
            recommendation = "Binary only — no ternary model"
        else:
            recommendation = "Not promising at steelmaking temps"

        values = [oxide, density, Tm, floats, solid,
                  binary_dG, binary_verdict,
                  tern_product, tern_dG, tern_verdict,
                  recommendation, tox]

        row_fill = white_fill if r_idx % 2 == 0 else alt_fill

        for c, val in enumerate(values, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.fill = row_fill

            # Binary dG column
            if c == 6:
                if isinstance(val, (int, float)):
                    cell.number_format = '+#,##0.0;-#,##0.0'
                    cell.font = red_font if val > 0 else green_font
                else:
                    cell.value = "TBD"
                    cell.fill = tbd_fill
                    cell.font = tbd_font
                cell.alignment = center
            # Binary verdict
            elif c == 7:
                cell.font = red_font if "CANNOT" in str(val) else green_font if "CAN" in str(val) else tbd_font
                cell.alignment = center
            # Ternary dG column
            elif c == 9:
                if isinstance(val, (int, float)):
                    cell.number_format = '+#,##0.0;-#,##0.0'
                    if val < -10:
                        cell.font = green_font
                        cell.fill = green_fill
                    elif val < 0:
                        cell.font = Font(color="006100", size=11)
                    else:
                        cell.font = red_font
                elif val is None:
                    cell.value = "TBD"
                    cell.fill = tbd_fill
                    cell.font = tbd_font
                cell.alignment = center
            # Ternary verdict
            elif c == 10:
                if "FAVORABLE" in str(val).upper() and "UN" not in str(val).upper():
                    cell.font = green_font
                elif "Unfavorable" in str(val):
                    cell.font = red_font
                else:
                    cell.font = normal_font
                cell.alignment = center
            # Overall recommendation
            elif c == 11:
                if "PROMISING" in str(val).upper():
                    cell.font = green_font
                    cell.fill = green_fill if row_fill == white_fill else green_fill
                elif "Not promising" in str(val):
                    cell.font = Font(color="9C0006", size=11)
                else:
                    cell.font = normal_font
                cell.alignment = left_wrap
            # Solid column
            elif c == 5 and val == "LIQUID":
                cell.font = Font(bold=True, color="9C0006", size=11)
                cell.alignment = center
            elif c in (1, 8, 12):
                cell.alignment = left_wrap
                cell.font = normal_font
            else:
                cell.alignment = center
                cell.font = normal_font

    last_row = len(SCREENING_ORDER) + 1
    apply_borders(ws, 1, last_row, 1, len(headers))

    add_notes(ws, last_row + 2, [
        "Binary: dG_rxn = dGf(Cu2O) - dGf(MOx). Positive = Cu cannot reduce oxide (all ceramics).",
        "Ternary: dG_rxn = G(CuMOy) - [G(Cu) + G(MOx) + G(O2)]. Negative = compound forms.",
        "\"Best Ternary Product\" is the compound with the most negative dG at 1527C.",
        "PROMISING oxides capture Cu into ternary compounds even though Cu cannot reduce them directly.",
        "This is the key insight: reduction is impossible, but compound formation may be favorable.",
    ])

    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 50


# ============================================================================
# Tab 6: Normalization Explained
# ============================================================================

def build_tab6_normalization(wb):
    """Educational tab explaining GM normalization."""
    ws = wb.create_sheet("Normalization Explained")

    # Title section
    ws.cell(row=1, column=1, value="HOW TC-PYTHON VALUES ARE NORMALIZED").font = bold_font
    ws.cell(row=3, column=1,
            value="Thermo-Calc reports GM = Gibbs energy per mole of ATOMS").font = normal_font
    ws.cell(row=4, column=1,
            value="We need: dGf per mole of O2 (for Ellingham comparison)").font = normal_font
    ws.cell(row=6, column=1,
            value="Step 1: GM (J/mol atoms) x atoms_per_formula = G per formula unit").font = normal_font
    ws.cell(row=7, column=1,
            value="Step 2: G per formula / oxide_per_O2 = dGf per mol O2").font = normal_font

    # Stoichiometry table
    ws.cell(row=9, column=1, value="STOICHIOMETRY TABLE").font = bold_font

    table_headers = ["Oxide", "Atoms", "Balanced Reaction", "oxide/O2"]
    for c, h in enumerate(table_headers, 1):
        cell = ws.cell(row=10, column=c, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 35
    ws.column_dimensions["D"].width = 10

    for r_idx, (oxide, atoms, reaction, ratio) in enumerate(NORM_TABLE):
        r = r_idx + 11
        row_fill = white_fill if r_idx % 2 == 0 else alt_fill
        for c, val in enumerate([oxide, atoms, reaction, ratio], 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = normal_font
            cell.fill = row_fill
            cell.alignment = center if c != 3 else left_wrap

    last_row = 10 + len(NORM_TABLE)
    apply_borders(ws, 10, last_row, 1, 4)


# ============================================================================
# Tab 7: Pricing & SDS
# ============================================================================

def build_tab7_pricing(wb):
    """Cost, CAS numbers, and safety data."""
    ws = wb.create_sheet("Pricing & SDS")

    headers = ["Oxide", "CAS", "Lab $/g", "Industrial $/g", "Sigma #", "Hazard"]
    col_widths = [10, 14, 10, 14, 10, 24]

    for c, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        ws.column_dimensions[get_column_letter(c)].width = w

    for r_idx, (oxide, cas, lab_price, ind_price, sigma_num) in enumerate(SDS_DATA):
        r = r_idx + 2
        hazard = OXIDE_PROPS.get(oxide, {}).get("toxicity", "")
        row_fill = white_fill if r_idx % 2 == 0 else alt_fill

        for c, val in enumerate([oxide, cas, lab_price, ind_price, sigma_num, hazard], 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = normal_font
            cell.fill = row_fill

            if c == 3:
                cell.number_format = '$#,##0.00'
                cell.alignment = center
            elif c == 4:
                cell.number_format = '$#,##0.0000'
                cell.alignment = center
            elif c in (1, 6):
                cell.alignment = left_wrap
            else:
                cell.alignment = center

            # Highlight HIGH hazard
            if c == 6 and "HIGH" in str(val).upper():
                cell.font = red_font

    last_row = len(SDS_DATA) + 1
    apply_borders(ws, 1, last_row, 1, len(headers))

    add_notes(ws, last_row + 2, [
        "Lab prices from Sigma-Aldrich (500g qty). Industrial prices are bulk estimates.",
        "CAS numbers for SDS lookup. Sigma catalog numbers for ordering.",
        "HIGH hazard oxides require additional safety review before experimental use.",
    ])

    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 30


# ============================================================================
# Main
# ============================================================================

def main():
    print("=" * 60)
    print("Building combined screening workbook...")
    print("=" * 60)

    df = load_binary_data()
    ternary_data = load_ternary_data()

    wb = openpyxl.Workbook()

    # Build all tabs
    build_tab1_formation_energies(wb, df)
    build_tab2_binary_screening(wb, df)
    build_tab3_ternary_balancing(wb)
    build_tab4_ternary_dG(wb, ternary_data)
    build_tab5_combined_screening(wb, df, ternary_data)
    build_tab6_normalization(wb)
    build_tab7_pricing(wb)

    wb.save(OUTPUT_FILE)
    print(f"\nSaved to: {OUTPUT_FILE}")
    print(f"Tabs: {[ws.title for ws in wb.worksheets]}")

    # Summary
    print(f"\nBinary data: {'loaded' if df is not None else 'NOT FOUND'}")
    print(f"Ternary data: {len(ternary_data)} rows" if ternary_data else "Ternary data: NOT YET AVAILABLE")
    if not ternary_data:
        print("  -> Run extract_ternary_reactions.py on VM, copy CSV, then re-run this script.")


if __name__ == "__main__":
    main()
