"""Build formatted screening table as .xlsx with colors and grouping."""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Oxide Screening"

# --- Style definitions (from CLAUDE.md IEEE standard) ---
header_fill = PatternFill(start_color="595959", end_color="595959", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
alt_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
tbd_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # light yellow for TBD
green_font = Font(color="006100", bold=True, size=11)
red_font = Font(color="9C0006", bold=True, size=11)
bold_font = Font(bold=True, size=11)
normal_font = Font(size=11)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)

thick = Side(style="medium")
thin = Side(style="thin")
thick_border = Border(left=thick, right=thick, top=thick, bottom=thick)

def apply_border(cell, left=thin, right=thin, top=thin, bottom=thin):
    cell.border = Border(left=left, right=right, top=top, bottom=bottom)

# --- Column headers ---
headers = [
    "Oxide",
    "Formula",
    "Density\n(g/cm3)",
    "Melting\nPoint (C)",
    "Floats on\nSteel?",
    "Solid at\n1527 C?",
    "dG_rxn @ 727 C\n(kJ/mol O2)",
    "dG_rxn @ 1227 C\n(kJ/mol O2)",
    "dG_rxn @ 1527 C\n(kJ/mol O2)",
    "Cu Can\nReduce?",
    "Toxicity",
    "Cost / Availability",
]

col_widths = [10, 10, 12, 12, 12, 12, 18, 18, 18, 12, 28, 25]

for c, (header, width) in enumerate(zip(headers, col_widths), 1):
    cell = ws.cell(row=1, column=c, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center
    ws.column_dimensions[get_column_letter(c)].width = width

# --- Data rows ---
# Group 1: Computed (white) — oxides with real data
# Group 2: TBD (light yellow) — need TC-Python

data = [
    # Group 1: COMPUTED FROM TCOX14
    ["MgO",   "MgO",    3.58, 2852, "Yes", "Yes", 796.0, 736.5, 642.9, "No", "Low", "Low - common refractory"],
    ["Al2O3", "Al2O3",  3.95, 2072, "Yes", "Yes", 716.7, 682.7, 644.8, "No", "Low", "Low - common ceramic"],
    ["TiO2",  "TiO2",   4.23, 1843, "Yes", "Yes", 569.4, 556.7, 531.3, "No", "Low", "Moderate"],
    ["SiO2",  "SiO2",   2.65, 1713, "Yes", "Yes", 539.4, 529.4, 502.7, "No", "Low (crystalline = inhalation hazard)", "Very low - sand"],
    ["FeO",   "FeO",    5.74, 1377, "Yes", "LIQUID", 220.4, 230.7, 213.8, "No", "Low", "Very low - mill scale"],
    ["CuO",   "CuO",    6.31, 1326, "Yes", "LIQUID", -58.8, -51.7, -41.0, "Trivial*", "Moderate - aquatic toxin", "Moderate"],
    # Group 2: TBD - NEED TC-PYTHON
    ["CaO",   "CaO",    3.34, 2613, "Yes", "Yes", "TBD", "TBD", "TBD", "TBD", "Low", "Very low - lime flux"],
    ["ZrO2",  "ZrO2",   5.68, 2715, "Yes", "Yes", "TBD", "TBD", "TBD", "TBD", "Low", "Moderate - refractory"],
    ["Cr2O3", "Cr2O3",  5.22, 2435, "Yes", "Yes", "TBD", "TBD", "TBD", "TBD", "Moderate - Cr(VI) risk", "Moderate"],
    ["MnO",   "MnO",    5.43, 1842, "Yes", "Yes", "TBD", "TBD", "TBD", "TBD", "Low", "Low - slag component"],
]

group1_end = 6   # rows 2-7 (first 6 oxides)
group2_start = 7  # rows 8-11 (TBD oxides)

for r, row_data in enumerate(data, 2):
    is_tbd = r > group1_end + 1  # group 2
    is_liquid = row_data[5] == "LIQUID"

    for c, val in enumerate(row_data, 1):
        cell = ws.cell(row=r, column=c)

        # Format dG values
        if c in (7, 8, 9) and isinstance(val, (int, float)):
            cell.value = val
            cell.number_format = '+#,##0.0;-#,##0.0'
            if val > 0:
                cell.font = red_font  # unfavorable = red
            else:
                cell.font = green_font  # favorable = green
        elif c in (7, 8, 9) and val == "TBD":
            cell.value = "TBD"
            cell.fill = tbd_fill
            cell.font = Font(italic=True, color="996600", size=11)
        elif c == 10 and val == "TBD":
            cell.value = "TBD"
            cell.fill = tbd_fill
            cell.font = Font(italic=True, color="996600", size=11)
        else:
            cell.value = val
            cell.font = normal_font

        # "LIQUID" in red
        if c == 6 and val == "LIQUID":
            cell.font = Font(bold=True, color="9C0006", size=11)

        # "No" in column 10 = bold
        if c == 10 and val == "No":
            cell.font = bold_font

        # Alignment
        if c in (1, 2, 11, 12):
            cell.alignment = left_wrap
        else:
            cell.alignment = center

    # Row fill: alternate by GROUP not by row
    # Group 1 = white, Group 2 = light gray
    if is_tbd:
        for c in range(1, 13):
            cell = ws.cell(row=r, column=c)
            if cell.fill == tbd_fill:
                pass  # keep yellow for TBD cells
            else:
                cell.fill = alt_fill

# --- Borders ---
# Thick outer frame
for r in range(1, len(data) + 2):
    for c in range(1, 13):
        cell = ws.cell(row=r, column=c)
        l = thick if c == 1 else thin
        ri = thick if c == 12 else thin
        t = thick if r == 1 else thin
        b = thick if r == len(data) + 1 else thin
        # Group divider between group 1 and group 2
        if r == group1_end + 1:  # last row of group 1
            b = thick
        if r == group1_end + 2:  # first row of group 2
            t = thick
        # Bottom of header
        if r == 1:
            b = thick
        apply_border(cell, left=l, right=ri, top=t, bottom=b)

# --- Group labels (merged cells below table) ---
note_row = len(data) + 3
ws.cell(row=note_row, column=1, value="Notes:").font = Font(bold=True, size=10)
ws.cell(row=note_row + 1, column=1, value="* CuO: Cu2O -> CuO is just further Cu oxidation, not useful for removal").font = Font(italic=True, size=10, color="666666")
ws.cell(row=note_row + 2, column=1, value="Molten steel density ~ 7.0 g/cm3 at 1600C. All oxides float.").font = Font(italic=True, size=10, color="666666")
ws.cell(row=note_row + 3, column=1, value="White rows = computed from TCOX14 data.  Gray rows = need TC-Python on ETS VM.").font = Font(italic=True, size=10, color="666666")
ws.cell(row=note_row + 4, column=1, value="Red dG values = UNFAVORABLE (Cu cannot reduce).  Green = favorable.").font = Font(italic=True, size=10, color="666666")
ws.cell(row=note_row + 5, column=1, value="dG_rxn = dG_f(Cu2O) - dG_f(MOx), per mol O2.  Positive = Cu CANNOT reduce oxide.").font = Font(italic=True, size=10, color="666666")

# Freeze top row
ws.freeze_panes = "A2"

# Row height
ws.row_dimensions[1].height = 35
for r in range(2, len(data) + 2):
    ws.row_dimensions[r].height = 22

out_path = "/Users/anthonydimascio/School/Spring2026/MSE-4381-Capstone/honda-calphad/screening/Oxide_Screening_Table.xlsx"
wb.save(out_path)
print(f"Saved to {out_path}")
