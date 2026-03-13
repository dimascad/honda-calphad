"""
Build formatted Excel workbook for ternary reaction screening results.

Tabs:
  1. Reaction Balancing — balanced reactions with atom counts (always available)
  2. Reaction dG — dG_rxn at key temperatures for each product (needs TC data)
  3. Screening Summary — final verdict table with color coding (needs TC data)

Run locally after compute_ternary_dG.py has been run.
Can also run without TC data to generate Tab 1 (reaction definitions only).
"""

import csv
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SCRIPT_DIR = Path(__file__).parent
CSV_RAW = SCRIPT_DIR.parent / "data" / "tcpython" / "raw" / "ternary_reaction_energies.csv"
CSV_PROCESSED = SCRIPT_DIR / "ternary_screening_results.csv"
OUTPUT_FILE = SCRIPT_DIR / "Ternary_Screening_Results.xlsx"

# ============================================================================
# Style definitions (IEEE standard from CLAUDE.md)
# ============================================================================
header_fill = PatternFill(start_color="595959", end_color="595959", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
alt_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
green_font = Font(color="006100", bold=True, size=11)
red_font = Font(color="9C0006", bold=True, size=11)
bold_font = Font(bold=True, size=11)
normal_font = Font(size=11)
small_font = Font(size=10, color="666666")
small_italic = Font(size=10, color="666666", italic=True)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)

thick = Side(style="medium")
thin = Side(style="thin")


def apply_borders(ws, row_start, row_end, col_start, col_end,
                  group_breaks=None):
    """Apply IEEE-style borders to a table region."""
    if group_breaks is None:
        group_breaks = set()

    for r in range(row_start, row_end + 1):
        for c in range(col_start, col_end + 1):
            cell = ws.cell(row=r, column=c)
            l = thick if c == col_start else thin
            ri = thick if c == col_end else thin
            t = thick if r == row_start else thin
            b = thick if r == row_end else thin
            # Header bottom
            if r == row_start:
                b = thick
            # Group breaks
            if r in group_breaks:
                b = thick
            if r - 1 in group_breaks and r != row_start:
                t = thick
            cell.border = Border(left=l, right=ri, top=t, bottom=b)


# ============================================================================
# Reaction definitions (same as extract_ternary_reactions.py)
# ============================================================================
REACTIONS = [
    # (product, name, reaction_str, oxide, n_Cu, n_oxide, oxide_formula, oxide_atoms, n_O2, prod_atoms, category)
    ("CuAl2O4",  "Copper aluminate spinel",       "Cu + Al2O3 + 0.5 O2 -> CuAl2O4",  "Al2O3", 1, 1,   "Al2O3", 5, 0.5,  7, "Spinel"),
    ("CuAlO2",   "Delafossite",                   "Cu + 0.5 Al2O3 + 0.25 O2 -> CuAlO2","Al2O3",1, 0.5, "Al2O3", 5, 0.25, 4, "Delafossite"),
    ("CuCr2O4",  "Copper chromite spinel",         "Cu + Cr2O3 + 0.5 O2 -> CuCr2O4",  "Cr2O3", 1, 1,   "Cr2O3", 5, 0.5,  7, "Spinel"),
    ("CuMn2O4",  "Copper manganite spinel",        "Cu + 2MnO + O2 -> CuMn2O4",        "MnO",   1, 2,   "MnO",   2, 1.0,  7, "Spinel"),
    ("CuFe2O4",  "Copper ferrite spinel",          "Cu + 2FeO + O2 -> CuFe2O4",        "FeO",   1, 2,   "FeO",   2, 1.0,  7, "Spinel"),
    ("CuCo2O4",  "Copper cobaltite spinel",        "Cu + 2CoO + O2 -> CuCo2O4",        "CoO",   1, 2,   "CoO",   2, 1.0,  7, "Spinel"),
    ("CuV2O6",   "Copper vanadate",                "Cu + V2O5 + 0.5 O2 -> CuV2O6",    "V2O5",  1, 1,   "V2O5",  7, 0.5,  9, "Vanadate"),
    ("Cu3V2O8",  "Copper orthovanadate",           "3Cu + V2O5 + 1.5 O2 -> Cu3V2O8",  "V2O5",  3, 1,   "V2O5",  7, 1.5, 13, "Vanadate"),
    ("CuTiO3",   "Copper titanate",                "Cu + TiO2 + 0.5 O2 -> CuTiO3",    "TiO2",  1, 1,   "TiO2",  3, 0.5,  5, "Titanate"),
    ("CuSiO3",   "Copper metasilicate",            "Cu + SiO2 + 0.5 O2 -> CuSiO3",    "SiO2",  1, 1,   "SiO2",  3, 0.5,  5, "Silicate"),
    ("Cu2SiO4",  "Copper orthosilicate",           "2Cu + SiO2 + O2 -> Cu2SiO4",       "SiO2",  2, 1,   "SiO2",  3, 1.0,  7, "Silicate"),
    ("CuMgO2",   "Copper magnesioxide (hyp.)",     "Cu + MgO + 0.5 O2 -> CuMgO2",     "MgO",   1, 1,   "MgO",   2, 0.5,  4, "Hypothetical"),
    ("CuCaO2",   "Copper calcioxide (hyp.)",       "Cu + CaO + 0.5 O2 -> CuCaO2",     "CaO",   1, 1,   "CaO",   2, 0.5,  4, "Hypothetical"),
    ("CuZrO3",   "Copper zirconate (hyp.)",        "Cu + ZrO2 + 0.5 O2 -> CuZrO3",    "ZrO2",  1, 1,   "ZrO2",  3, 0.5,  5, "Hypothetical"),
    ("CuNiO2",   "Copper nickelate",               "Cu + NiO + 0.5 O2 -> CuNiO2",     "NiO",   1, 1,   "NiO",   2, 0.5,  4, "Compound"),
    ("CuLaO2",   "Copper lanthanum oxide",         "Cu + 0.5 La2O3 + 0.25 O2 -> CuLaO2","La2O3",1, 0.5,"La2O3", 5, 0.25, 4, "Compound"),
    ("CuCeO3",   "Copper cerate (hyp.)",           "Cu + CeO2 + 0.5 O2 -> CuCeO3",    "CeO2",  1, 1,   "CeO2",  3, 0.5,  5, "Hypothetical"),
    ("CuB2O4",   "Copper borate",                  "Cu + B2O3 + 0.5 O2 -> CuB2O4",    "B2O3",  1, 1,   "B2O3",  5, 0.5,  7, "Borate"),
]


def build_tab1_reaction_balancing(wb):
    """Tab 1: Reaction balancing — shows each reaction with atom counts."""
    ws = wb.active
    ws.title = "Reaction Balancing"

    headers = [
        "Product",
        "Name",
        "Category",
        "Balanced Reaction",
        "Cu\n(reactant)",
        "Oxide\nFormula",
        "n(oxide)\n(mol)",
        "Oxide\nAtoms",
        "n(O2)\n(mol)",
        "O atoms\nfrom O2",
        "Total\nReactant\nAtoms",
        "Product\nAtoms",
        "Balanced?",
    ]
    col_widths = [14, 28, 14, 40, 10, 10, 10, 10, 10, 10, 12, 12, 10]

    for c, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        ws.column_dimensions[get_column_letter(c)].width = w

    # Group categories for alternating shading
    categories_seen = []
    for rxn in REACTIONS:
        cat = rxn[10]
        if not categories_seen or categories_seen[-1] != cat:
            categories_seen.append(cat)

    cat_group_idx = {}
    for i, cat in enumerate(categories_seen):
        cat_group_idx[cat] = i

    group_break_rows = set()
    prev_cat = None

    for r, rxn in enumerate(REACTIONS, 2):
        product, name, reaction, oxide, n_Cu, n_oxide, ox_formula, ox_atoms, n_O2, prod_atoms, category = rxn

        o_from_O2 = n_O2 * 2
        total_reactant = n_Cu + n_oxide * ox_atoms + o_from_O2
        balanced = "YES" if abs(total_reactant - prod_atoms) < 0.01 else "NO"

        values = [product, name, category, reaction,
                  n_Cu, ox_formula, n_oxide, ox_atoms, n_O2, o_from_O2,
                  total_reactant, prod_atoms, balanced]

        # Track group breaks
        if prev_cat is not None and category != prev_cat:
            group_break_rows.add(r - 1)  # bottom of previous group
        prev_cat = category

        # Alternating fill by category group
        group_num = cat_group_idx.get(category, 0)
        row_fill = alt_fill if group_num % 2 == 1 else white_fill

        for c, val in enumerate(values, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = normal_font
            cell.fill = row_fill
            if c in (1, 2, 3, 4, 6):
                cell.alignment = left_wrap
            else:
                cell.alignment = center

            # Color the balanced check
            if c == 13:
                if val == "YES":
                    cell.font = green_font
                else:
                    cell.font = red_font

    # Borders
    apply_borders(ws, 1, len(REACTIONS) + 1, 1, 13, group_break_rows)

    # Notes
    note_row = len(REACTIONS) + 3
    ws.cell(row=note_row, column=1, value="Notes:").font = bold_font
    notes = [
        "All reactions require oxygen. In steelmaking, O2 comes from dissolved oxygen in the melt or from the atmosphere.",
        "n(oxide) = moles of binary oxide consumed per mole of ternary product formed.",
        "Oxide Atoms = atoms per formula unit of the binary oxide (e.g., Al2O3 = 5 atoms).",
        "Total Reactant Atoms must equal Product Atoms for the reaction to be balanced.",
        "Hypothetical compounds have no known crystal structure; TC-Python will determine if they are stable.",
    ]
    for i, note in enumerate(notes, 1):
        ws.cell(row=note_row + i, column=1, value=note).font = small_italic

    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 40
    return ws


def build_tab2_reaction_dG(wb, raw_data):
    """Tab 2: Reaction dG at key temperatures."""
    ws = wb.create_sheet("Reaction dG")

    TEMPS_K = [1000, 1500, 1800]

    headers = ["Product", "Name", "Oxide", "Reaction"]
    for T in TEMPS_K:
        headers.append(f"dG_rxn\n@ {T-273}C\n(kJ)")
    headers.extend(["Ternary Phase\n@ 1527C", "Stable Phases\n@ 1527C"])
    col_widths = [14, 28, 10, 40, 14, 14, 14, 20, 40]

    for c, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        ws.column_dimensions[get_column_letter(c)].width = w

    # Organize raw data by product
    by_product = {}
    for row in raw_data:
        key = row["product"]
        if key not in by_product:
            by_product[key] = []
        by_product[key].append(row)

    import numpy as np

    group_break_rows = set()
    prev_cat = None

    for r_idx, rxn_def in enumerate(REACTIONS):
        r = r_idx + 2
        product = rxn_def[0]
        name = rxn_def[1]
        oxide = rxn_def[3]
        reaction = rxn_def[2]
        category = rxn_def[10]

        if prev_cat is not None and category != prev_cat:
            group_break_rows.add(r - 1)
        prev_cat = category

        row_fill = white_fill if r_idx % 2 == 0 else alt_fill
        values = [product, name, oxide, reaction]

        prod_rows = by_product.get(product, [])
        T_arr = []
        dG_arr = []
        for pr in prod_rows:
            try:
                T_arr.append(float(pr["T_K"]))
                dG_str = pr.get("dG_rxn_system_kJ", "").strip()
                dG_arr.append(float(dG_str) if dG_str else None)
            except (ValueError, KeyError):
                pass

        # Get dG at each key temperature
        for T_target in TEMPS_K:
            if T_arr:
                np_T = np.array([t for t, d in zip(T_arr, dG_arr) if d is not None])
                np_dG = np.array([d for d in dG_arr if d is not None])
                if len(np_T) > 0:
                    idx = int(np.argmin(np.abs(np_T - T_target)))
                    if abs(np_T[idx] - T_target) < 30:
                        values.append(round(np_dG[idx], 1))
                    else:
                        values.append("No data")
                else:
                    values.append("No data")
            else:
                values.append("TBD")

        # Ternary phase at 1800K
        rows_1800 = [pr for pr in prod_rows
                     if pr.get("T_K") and abs(float(pr["T_K"]) - 1800) < 30]
        if rows_1800:
            values.append(rows_1800[0].get("ternary_phase_found", ""))
            values.append(rows_1800[0].get("stable_phases", ""))
        else:
            values.append("")
            values.append("TBD — run extract_ternary_reactions.py")

        for c, val in enumerate(values, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.fill = row_fill

            # Color dG values
            if c in (5, 6, 7) and isinstance(val, (int, float)):
                cell.number_format = '+#,##0.0;-#,##0.0'
                if val < -10:
                    cell.font = green_font
                    cell.fill = green_fill
                elif val < 0:
                    cell.font = Font(color="006100", size=11)
                else:
                    cell.font = red_font
            elif c in (5, 6, 7):
                cell.font = Font(italic=True, color="996600", size=11)
            else:
                cell.font = normal_font

            if c in (1, 2, 4, 8, 9):
                cell.alignment = left_wrap
            else:
                cell.alignment = center

    apply_borders(ws, 1, len(REACTIONS) + 1, 1, len(headers), group_break_rows)

    # Formula explanation
    note_row = len(REACTIONS) + 3
    ws.cell(row=note_row, column=1, value="Calculation:").font = bold_font
    formulas = [
        "dG_rxn = G(products) - G(reactants)",
        "G(products) = GM_system(ternary composition) x atoms_per_formula",
        "G(reactants) = n_Cu x G(Cu metal) + n_oxide x oxide_atoms x G(binary oxide) + n_O2 x G(O2 gas)",
        "Negative dG = FAVORABLE (oxide captures copper into ternary compound)",
        "Positive dG = UNFAVORABLE (ternary compound does not form)",
        "Green cells = favorable reactions.  Red = unfavorable.",
    ]
    for i, f in enumerate(formulas, 1):
        ws.cell(row=note_row + i, column=1, value=f).font = small_italic

    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 50


def build_tab3_screening_summary(wb, raw_data):
    """Tab 3: Screening summary — one row per oxide with best ternary reaction."""
    ws = wb.create_sheet("Screening Summary")

    headers = [
        "Oxide",
        "Best Ternary\nProduct",
        "Type",
        "dG_rxn @ 727C\n(kJ/mol)",
        "dG_rxn @ 1227C\n(kJ/mol)",
        "dG_rxn @ 1527C\n(kJ/mol)",
        "Ternary Phase\nFound?",
        "Verdict",
        "Notes",
    ]
    col_widths = [10, 16, 14, 16, 16, 16, 16, 22, 35]

    for c, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        ws.column_dimensions[get_column_letter(c)].width = w

    # Group by oxide, pick best (most negative dG at 1800K) product
    import numpy as np

    oxide_best = {}
    by_product = {}
    for row in raw_data:
        key = row["product"]
        if key not in by_product:
            by_product[key] = []
        by_product[key].append(row)

    TEMPS_K = [1000, 1500, 1800]

    for rxn_def in REACTIONS:
        product = rxn_def[0]
        oxide = rxn_def[3]
        category = rxn_def[10]

        prod_rows = by_product.get(product, [])
        # Get dG at 1800K
        dG_1800 = None
        dG_values = {}
        for T_target in TEMPS_K:
            for pr in prod_rows:
                try:
                    T = float(pr["T_K"])
                    if abs(T - T_target) < 30:
                        dG_str = pr.get("dG_rxn_system_kJ", "").strip()
                        if dG_str:
                            dG_values[T_target] = float(dG_str)
                            if T_target == 1800:
                                dG_1800 = float(dG_str)
                except (ValueError, KeyError):
                    pass

        # Phase info at 1800K
        rows_1800 = [pr for pr in prod_rows
                     if pr.get("T_K") and abs(float(pr["T_K"]) - 1800) < 30]
        phase_found = ""
        if rows_1800:
            phase_found = rows_1800[0].get("ternary_phase_found", "")

        if oxide not in oxide_best or (dG_1800 is not None and
                (oxide_best[oxide]["dG_1800"] is None or dG_1800 < oxide_best[oxide]["dG_1800"])):
            oxide_best[oxide] = {
                "product": product,
                "category": category,
                "dG_values": dG_values,
                "dG_1800": dG_1800,
                "phase_found": phase_found,
            }

    # Write rows ordered by dG_1800
    oxide_order = sorted(oxide_best.keys(),
                         key=lambda x: oxide_best[x]["dG_1800"] if oxide_best[x]["dG_1800"] is not None else 9999)

    for r_idx, oxide in enumerate(oxide_order):
        r = r_idx + 2
        info = oxide_best[oxide]
        row_fill = white_fill if r_idx % 2 == 0 else alt_fill

        dG_1800 = info["dG_1800"]
        if dG_1800 is not None:
            if dG_1800 < -10:
                verdict = "FAVORABLE — Cu captured"
            elif dG_1800 < 0:
                verdict = "Marginally favorable"
            elif dG_1800 < 10:
                verdict = "Borderline"
            else:
                verdict = "Unfavorable"
        else:
            verdict = "No data"

        notes = ""
        if info["phase_found"]:
            notes = f"TC found: {info['phase_found']}"
        elif dG_1800 is not None and dG_1800 < 0:
            notes = "Favorable but no named phase found — check TC output"

        values = [
            oxide,
            info["product"],
            info["category"],
        ]
        for T in TEMPS_K:
            v = info["dG_values"].get(T)
            values.append(round(v, 1) if v is not None else "TBD")
        values.extend([info["phase_found"] or "--", verdict, notes])

        for c, val in enumerate(values, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.fill = row_fill

            if c in (4, 5, 6) and isinstance(val, (int, float)):
                cell.number_format = '+#,##0.0;-#,##0.0'
                if val < -10:
                    cell.font = green_font
                    cell.fill = green_fill
                elif val < 0:
                    cell.font = Font(color="006100", size=11)
                else:
                    cell.font = red_font
            elif c == 8:
                if "FAVORABLE" in str(verdict).upper() and "UN" not in str(verdict).upper():
                    cell.font = green_font
                elif "Unfavorable" in str(verdict):
                    cell.font = red_font
                else:
                    cell.font = normal_font
            elif c in (4, 5, 6):
                cell.font = Font(italic=True, color="996600", size=11)
            else:
                cell.font = normal_font

            if c in (1, 2, 3, 7, 8, 9):
                cell.alignment = left_wrap
            else:
                cell.alignment = center

    apply_borders(ws, 1, len(oxide_order) + 1, 1, len(headers))

    note_row = len(oxide_order) + 3
    ws.cell(row=note_row, column=1, value="Notes:").font = bold_font
    notes = [
        "Ranked by most favorable (most negative) dG_rxn at steelmaking temperature (1527C).",
        "For oxides with multiple possible ternary products, only the BEST (most negative dG) is shown.",
        "\"Hypothetical\" compounds have no known crystal structure; they may not exist.",
        "Green = favorable for copper capture.  Red = unfavorable.",
    ]
    for i, n in enumerate(notes, 1):
        ws.cell(row=note_row + i, column=1, value=n).font = small_italic

    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 40


def main():
    wb = openpyxl.Workbook()

    # Tab 1: Reaction Balancing (always available — no TC data needed)
    build_tab1_reaction_balancing(wb)

    # Try to load TC data for Tabs 2 and 3
    raw_data = []
    if CSV_RAW.exists():
        with open(CSV_RAW) as f:
            raw_data = list(csv.DictReader(f))
        print(f"Loaded {len(raw_data)} rows from {CSV_RAW}")
    else:
        print(f"No TC data found at {CSV_RAW}")
        print("Tab 1 (Reaction Balancing) will be generated.")
        print("Tabs 2-3 will show TBD placeholders.")
        print("Run extract_ternary_reactions.py on the VM first.")

    build_tab2_reaction_dG(wb, raw_data)
    build_tab3_screening_summary(wb, raw_data)

    wb.save(OUTPUT_FILE)
    print(f"\nSaved to: {OUTPUT_FILE}")
    print(f"Tabs: {[ws.title for ws in wb.worksheets]}")


if __name__ == "__main__":
    main()
