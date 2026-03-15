"""
Build Cu_Removal_Unified.xlsx — ONE workbook, fully traceable.

Tab layout (narrative order, 3-tier color coding):

  BLUE  — Analysis (formula-driven, what Honda/Zhang see)
    1. Dashboard              — project story, navigation card, 4 embedded plots
    2. Screening_Summary      — binary + ternary verdicts, tier classification
    3. Mass_Balance           — experiment design: oxide dropdown, particle sweep, contact time
    4. Activity_and_Slag      — dilute Cu penalty + slag basicity effects
    5. CuFe2O4_DeepDive       — Cu-capture vs Fe-oxidation decomposition
    6. Normalization          — full stoichiometry tables (17 binary + 18 ternary) + 2 worked examples

  GREEN — Results (derived CSVs — still "answers", not raw TC-Python)
    7.  Screening_Table       — one row per oxide, binary dG + verdict
    8.  Ternary_Verdicts      — one row per ternary product, dG + verdict
    9.  Activity_Corrected    — activity-corrected dG across temperatures

  GRAY  — Raw TC-Python Output (the receipt)
    10. OX_Gibbs              — binary oxide Gibbs energies (18 oxides x 31 temps)
    11. Ternary_Rxns          — ternary reaction energies (414 rows)
    12. dG_Top6               — fine-resolution dG vs T for top 6 (270 rows)
    13. CuFe2O4_Raw           — CuFe2O4 decomposition raw data (23 rows)
    14. Cu_Activity           — Cu activity sweep at 1800K (80 rows)
    15. Slag_Effects          — slag basicity effects on a_Cu (30 rows)
    16. Phase_Map             — ternary composition phase mapping (1107 rows)
"""

import csv
from pathlib import Path

import openpyxl
from openpyxl.drawing.image import Image as XlImage
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers,
)
from openpyxl.formatting.rule import CellIsRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.axis import ChartLines
from openpyxl.drawing.line import LineProperties
from openpyxl.chart.shapes import GraphicalProperties

# ── Matplotlib-consistent colors and dash styles for Excel charts ─────
# Matches compute_activity_corrected_dG.py COLORS and LINE_STYLES dicts.
# matplotlib "-" → "solid", "--" → "dash", ":" → "dot", "-." → "dashDot"
PRODUCT_CHART_STYLE = {
    #             (hex color,  openpyxl dash style)
    "CuFe2O4":   ("0077BB", "solid"),
    "CuMn2O4":   ("AA3377", "dot"),
    "CuB2O4":    ("EE3377", "solid"),
    "CuV2O6":    ("CC6600", "dashDot"),
    "CuAl2O4":   ("BBBBBB", "dash"),
    "Cu3V2O8":   ("EE7733", "dash"),
    "Cu2SiO4":   ("009988", "dashDot"),
}

# ── paths ──────────────────────────────────────────────────────────────
SCRIPT_DIR = Path(__file__).parent
DATA_DIR = SCRIPT_DIR.parent / "data" / "tcpython" / "raw"
PROC_DIR = SCRIPT_DIR.parent / "data" / "tcpython" / "processed"
FIG_DIR  = SCRIPT_DIR.parent / "figures"
OUTPUT_FILE = SCRIPT_DIR / "Cu_Removal_Unified.xlsx"

CSV_FILES = {
    "ox_gibbs":     DATA_DIR / "oxide_gibbs_energies.csv",
    "ternary_rxns": DATA_DIR / "ternary_reaction_energies.csv",
    "dG_top6":      DATA_DIR / "dG_vs_T_top6.csv",
    "cufe2o4":      DATA_DIR / "cufe2o4_alternative_reaction.csv",
    "cu_activity":  DATA_DIR / "cu_activity_vs_oxide.csv",
    "slag":         DATA_DIR / "slag_composition_effects.csv",
    "phase_map":    DATA_DIR / "ternary_phase_map_1800K.csv",
    "corrected":    PROC_DIR / "activity_corrected_dG.csv",
    "screening":    SCRIPT_DIR / "screening_table.csv",
    "ternary_v":    SCRIPT_DIR / "ternary_screening_results.csv",
}

FIGURES = {
    "dG_vs_T":    FIG_DIR / "dG_vs_T_top6.png",
    "activity":   FIG_DIR / "dG_corrected_comparison.png",
    "slag":       FIG_DIR / "slag_basicity_vs_aCu.png",
    "cufe2o4":    FIG_DIR / "cufe2o4_decomposition.png",
}

# ── IEEE formatting constants ──────────────────────────────────────────
HEADER_FILL = PatternFill(start_color="595959", end_color="595959",
                          fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
WHITE_FILL  = PatternFill(start_color="FFFFFF", end_color="FFFFFF",
                          fill_type="solid")
GRAY_FILL   = PatternFill(start_color="E0E0E0", end_color="E0E0E0",
                          fill_type="solid")
GREEN_FILL  = PatternFill(start_color="C6EFCE", end_color="C6EFCE",
                          fill_type="solid")
RED_FILL    = PatternFill(start_color="FFC7CE", end_color="FFC7CE",
                          fill_type="solid")
# Dashboard card background
CARD_FILL   = PatternFill(start_color="F2F2F2", end_color="F2F2F2",
                          fill_type="solid")

THICK = Side(style="medium")
THIN  = Side(style="thin")

# Dropdown cell style — bright orange-gold fill so it pops as "click me"
DROPDOWN_FILL = PatternFill(start_color="FFC000", end_color="FFC000",
                             fill_type="solid")
DROPDOWN_FONT = Font(bold=True, size=11, color="000000")
DROPDOWN_BORDER = Border(
    top=Side(style="medium", color="B8860B"),
    bottom=Side(style="medium", color="B8860B"),
    left=Side(style="medium", color="B8860B"),
    right=Side(style="medium", color="B8860B"))

# Tab colors (3-tier)
BLUE_TAB  = "4472C4"   # analysis
GREEN_TAB = "70AD47"   # derived results
GRAY_TAB  = "A5A5A5"   # raw TC-Python
DASH_TAB  = "2F5496"   # dashboard (darker blue)

# Number of intro-card rows added to raw/results tabs.
# All cross-tab formula references must add this offset.
INTRO_ROWS = 2

# ── Normalization stoichiometry ───────────────────────────────────────
NORM_TABLE = [
    # (oxide, atoms_per_formula, balanced_reaction, oxide_per_O2)
    ("Cu2O",  3, "4Cu + O2 -> 2Cu2O",              2),
    ("Al2O3", 5, "4/3 Al + O2 -> 2/3 Al2O3",       0.667),
    ("MgO",   2, "2Mg + O2 -> 2MgO",               2),
    ("SiO2",  3, "Si + O2 -> SiO2",                 1),
    ("TiO2",  3, "Ti + O2 -> TiO2",                 1),
    ("FeO",   2, "2Fe + O2 -> 2FeO",                2),
    ("CaO",   2, "2Ca + O2 -> 2CaO",                2),
    ("ZrO2",  3, "Zr + O2 -> ZrO2",                 1),
    ("Cr2O3", 5, "4/3 Cr + O2 -> 2/3 Cr2O3",       0.667),
    ("MnO",   2, "2Mn + O2 -> 2MnO",                2),
    ("V2O5",  7, "4/5 V + O2 -> 2/5 V2O5",          0.4),
    ("B2O3",  5, "4/3 B + O2 -> 2/3 B2O3",          0.667),
    ("La2O3", 5, "4/3 La + O2 -> 2/3 La2O3",        0.667),
    ("CeO2",  3, "Ce + O2 -> CeO2",                 1),
    ("NiO",   2, "2Ni + O2 -> 2NiO",                2),
    ("CoO",   2, "2Co + O2 -> 2CoO",                2),
    ("PbO",   2, "2Pb + O2 -> 2PbO",                2),
]

TERNARY_REACTIONS = [
    # (product, name, reaction, oxide, n_Cu, n_oxide, n_O2, prod_atoms, category)
    ("CuAl2O4",  "Cu aluminate spinel",   "Cu + Al2O3 + 0.5O2 -> CuAl2O4",    "Al2O3", 1, 1,   0.5,  7, "Spinel"),
    ("CuAlO2",   "Delafossite",           "Cu + 0.5Al2O3 + 0.25O2 -> CuAlO2", "Al2O3", 1, 0.5, 0.25, 4, "Delafossite"),
    ("CuCr2O4",  "Cu chromite spinel",    "Cu + Cr2O3 + 0.5O2 -> CuCr2O4",    "Cr2O3", 1, 1,   0.5,  7, "Spinel"),
    ("CuMn2O4",  "Cu manganite spinel",   "Cu + 2MnO + O2 -> CuMn2O4",        "MnO",   1, 2,   1.0,  7, "Spinel"),
    ("CuFe2O4",  "Cu ferrite spinel",     "Cu + 2FeO + O2 -> CuFe2O4",        "FeO",   1, 2,   1.0,  7, "Spinel"),
    ("CuCo2O4",  "Cu cobaltite spinel",   "Cu + 2CoO + O2 -> CuCo2O4",        "CoO",   1, 2,   1.0,  7, "Spinel"),
    ("CuV2O6",   "Copper vanadate",       "Cu + V2O5 + 0.5O2 -> CuV2O6",      "V2O5",  1, 1,   0.5,  9, "Vanadate"),
    ("Cu3V2O8",  "Cu orthovanadate",      "3Cu + V2O5 + 1.5O2 -> Cu3V2O8",    "V2O5",  3, 1,   1.5, 13, "Vanadate"),
    ("CuTiO3",   "Copper titanate",       "Cu + TiO2 + 0.5O2 -> CuTiO3",      "TiO2",  1, 1,   0.5,  5, "Titanate"),
    ("CuSiO3",   "Cu metasilicate",       "Cu + SiO2 + 0.5O2 -> CuSiO3",      "SiO2",  1, 1,   0.5,  5, "Silicate"),
    ("Cu2SiO4",  "Cu orthosilicate",      "2Cu + SiO2 + O2 -> Cu2SiO4",       "SiO2",  2, 1,   1.0,  7, "Silicate"),
    ("CuMgO2",   "Cu magnesioxide (hyp)", "Cu + MgO + 0.5O2 -> CuMgO2",       "MgO",   1, 1,   0.5,  4, "Hypothetical"),
    ("CuCaO2",   "Cu calcioxide (hyp)",   "Cu + CaO + 0.5O2 -> CuCaO2",       "CaO",   1, 1,   0.5,  4, "Hypothetical"),
    ("CuZrO3",   "Cu zirconate (hyp)",    "Cu + ZrO2 + 0.5O2 -> CuZrO3",      "ZrO2",  1, 1,   0.5,  5, "Hypothetical"),
    ("CuNiO2",   "Copper nickelate",      "Cu + NiO + 0.5O2 -> CuNiO2",       "NiO",   1, 1,   0.5,  4, "Compound"),
    ("CuLaO2",   "Cu lanthanum oxide",    "Cu + 0.5La2O3 + 0.25O2 -> CuLaO2", "La2O3", 1, 0.5, 0.25, 4, "Compound"),
    ("CuCeO3",   "Cu cerate (hyp)",       "Cu + CeO2 + 0.5O2 -> CuCeO3",      "CeO2",  1, 1,   0.5,  5, "Hypothetical"),
    ("CuB2O4",   "Copper borate",         "Cu + B2O3 + 0.5O2 -> CuB2O4",      "B2O3",  1, 1,   0.5,  7, "Borate"),
]

# Toxicity highlighting
TOXICITY_KEYWORDS = {"HIGH", "CARC", "IARC", "NEUROTOXIN"}
TOXICITY_FONT = Font(bold=True, color="CC0000", size=10)
TOXICITY_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE",
                            fill_type="solid")


def _highlight_toxicity(ws, col_idx, start_row, end_row):
    """Apply red bold font to cells containing HIGH toxicity keywords."""
    for r in range(start_row, end_row + 1):
        cell = ws.cell(row=r, column=col_idx)
        val = str(cell.value or "").upper()
        if any(kw in val for kw in TOXICITY_KEYWORDS):
            cell.font = TOXICITY_FONT


# ── helpers ────────────────────────────────────────────────────────────
def load_csv(path):
    """Return (fieldnames, list[dict])."""
    if not path.exists():
        return None, []
    with open(path, encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        fieldnames = list(reader.fieldnames)
        rows = list(reader)
    return fieldnames, rows


def _try_float(v):
    if v is None:
        return v
    try:
        return float(v)
    except (TypeError, ValueError):
        return v


def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)


def apply_row_fill(ws, row, ncols, fill):
    for c in range(1, ncols + 1):
        ws.cell(row=row, column=c).fill = fill


def alt_shading(ws, start_row, end_row, ncols, group_size=1):
    for r in range(start_row, end_row + 1):
        grp = (r - start_row) // group_size
        fill = WHITE_FILL if grp % 2 == 0 else GRAY_FILL
        apply_row_fill(ws, r, ncols, fill)


def group_shading(ws, start_row, end_row, ncols, group_col=1):
    grp_idx = 0
    prev_val = None
    for r in range(start_row, end_row + 1):
        val = ws.cell(row=r, column=group_col).value
        if val != prev_val:
            grp_idx += 1
            prev_val = val
        fill = WHITE_FILL if grp_idx % 2 == 1 else GRAY_FILL
        apply_row_fill(ws, r, ncols, fill)


def apply_borders(ws, start_row, end_row, ncols):
    for r in range(start_row, end_row + 1):
        for c in range(1, ncols + 1):
            top = THICK if r == start_row else THIN
            bot = THICK if r == end_row else THIN
            left = THICK if c == 1 else THIN
            right = THICK if c == ncols else THIN
            if r == start_row:
                bot = THICK
            ws.cell(row=r, column=c).border = Border(
                top=top, bottom=bot, left=left, right=right
            )


def auto_width(ws, ncols, min_w=8, max_w=30):
    for c in range(1, ncols + 1):
        best = min_w
        for row_cells in ws.iter_rows(min_col=c, max_col=c, values_only=False):
            for ce in row_cells:
                if ce.value is not None:
                    best = max(best, min(len(str(ce.value)) + 2, max_w))
        ws.column_dimensions[get_column_letter(c)].width = min(best, max_w)


def center_all(ws, start_row, end_row, ncols):
    for r in range(start_row, end_row + 1):
        for c in range(1, ncols + 1):
            ws.cell(row=r, column=c).alignment = Alignment(horizontal="center")


def fmt_number(ws, row, col, fmt_str):
    ws.cell(row=row, column=col).number_format = fmt_str


def freeze_and_color(ws, tab_color, freeze_row=2):
    ws.sheet_properties.tabColor = tab_color
    ws.freeze_panes = f"A{freeze_row}"


def _write_card_line(ws, row, col_start, col_end, text,
                     font=None, fill=CARD_FILL, merge=True):
    """Write a single-row 'card' line across columns with fill."""
    cell = ws.cell(row=row, column=col_start, value=text)
    if font:
        cell.font = font
    for c in range(col_start, col_end + 1):
        ws.cell(row=row, column=c).fill = fill
    if merge and col_end > col_start:
        ws.merge_cells(start_row=row, start_column=col_start,
                       end_row=row, end_column=col_end)


def _embed_image(ws, img_path, anchor_cell, width_px=500):
    """Embed a PNG image at anchor_cell, scaled to width_px."""
    if not img_path.exists():
        ws.cell(row=int(anchor_cell[1:]), column=1,
                value=f"[Image not found: {img_path.name}]")
        return 0
    img = XlImage(str(img_path))
    # Scale proportionally
    aspect = img.height / img.width
    img.width = width_px
    img.height = int(width_px * aspect)
    ws.add_image(img, anchor_cell)
    # Return approximate row height consumed (assuming ~15px per row)
    return max(1, img.height // 15)


# =====================================================================
# RAW DATA TABS (7 tabs — direct TC-Python CSV dumps)
# =====================================================================
def _build_raw_tab(wb, tab_name, csv_key, tab_color=GRAY_TAB, intro=None):
    """Build a CSV-backed tab with optional intro card lines.

    intro: list of (text, Font) tuples for 1-2 card lines above the header.
    """
    fields, rows = load_csv(CSV_FILES[csv_key])
    if not rows:
        ws = wb.create_sheet(tab_name)
        ws.cell(row=1, column=1, value=f"Data not available ({csv_key})")
        freeze_and_color(ws, tab_color)
        return ws, 0

    ws = wb.create_sheet(tab_name)
    ncols = len(fields)

    offset = 0
    if intro:
        merge_end = min(ncols, 10)
        for idx, (text, font) in enumerate(intro):
            _write_card_line(ws, idx + 1, 1, merge_end, text, font=font)
        offset = len(intro)

    hdr_row = offset + 1
    for c, h in enumerate(fields, 1):
        ws.cell(row=hdr_row, column=c, value=h)
    style_header(ws, hdr_row, ncols)

    for i, r in enumerate(rows):
        row = hdr_row + 1 + i
        for c, h in enumerate(fields, 1):
            ws.cell(row=row, column=c, value=_try_float(r.get(h, "")))

    last_row = hdr_row + len(rows)
    alt_shading(ws, hdr_row + 1, last_row, ncols)
    center_all(ws, hdr_row, last_row, ncols)
    apply_borders(ws, hdr_row, last_row, ncols)
    auto_width(ws, ncols)
    freeze_and_color(ws, tab_color, freeze_row=hdr_row + 1)

    print(f"  {tab_name}: {len(rows)} rows, {ncols} cols")
    return ws, len(rows)


def _intro(title, detail):
    """Return a 2-line intro card list for _build_raw_tab."""
    return [
        (title, Font(bold=True, size=11, color="2F5496")),
        (detail, Font(size=9, italic=True, color="666666")),
    ]


def build_raw_tabs(wb):
    """Build 8 raw/backing data tabs (gray) with intro cards."""
    _build_raw_tab(wb, "Ternary_Verdicts", "ternary_v",
        tab_color=GRAY_TAB, intro=_intro(
        "TERNARY VERDICTS \u2014 One row per Cu-M-O product "
        "(backing data for Screening_Summary)",
        "\u0394G < 0 means oxide captures Cu into a ternary compound. "
        "Col I shows equilibrium phases at 1527\u00b0C (all melt into "
        "IONIC_LIQ at steelmaking temp)."))
    _build_raw_tab(wb, "OX_Gibbs", "ox_gibbs", intro=_intro(
        "RAW TC-PYTHON \u2014 Binary oxide Gibbs energies (18 oxides, TCOX14+SSUB3)",
        "GM values are per mole of ATOMS. See Normalization tab for conversion "
        "to kJ/mol O\u2082. Temperatures: 500-2000 K in 50 K steps."))
    _build_raw_tab(wb, "Ternary_Rxns", "ternary_rxns", intro=_intro(
        "RAW TC-PYTHON \u2014 Ternary reaction energies (16 systems, TCOX14)",
        "414 rows: 23 temperatures (800-1900 K, 50 K steps) \u00d7 18 products. "
        "dG_rxn_kJ = total reaction driving force."))
    _build_raw_tab(wb, "dG_Top6", "dG_top6", intro=_intro(
        "RAW TC-PYTHON \u2014 Fine-resolution \u0394G vs T for top 6 candidates",
        "25 K steps for higher resolution than the 50 K screening grid. "
        "Source for the dG_vs_T_top6.png figure on the Dashboard."))
    _build_raw_tab(wb, "CuFe2O4_Raw", "cufe2o4", intro=_intro(
        "RAW TC-PYTHON \u2014 CuFe\u2082O\u2084 decomposition components",
        "Splits total reaction into Cu-capture + Fe-oxidation. "
        "See CuFe2O4_DeepDive tab for analysis with formula cells."))
    _build_raw_tab(wb, "Cu_Activity", "cu_activity", intro=_intro(
        "RAW TC-PYTHON \u2014 Cu activity sweep at 1800 K (4 systems)",
        "a_Cu is invariant in 2-phase fields (Gibbs Phase Rule). "
        "TCOX14 has no separate metallic Cu phase; all Cu enters IONIC_LIQ."))
    _build_raw_tab(wb, "Slag_Effects", "slag", intro=_intro(
        "RAW TC-PYTHON \u2014 Slag basicity effects on Cu activity",
        "MnO:SiO\u2082 and Al\u2082O\u2083:SiO\u2082 ratio variations "
        "at 1800 K. See Activity_and_Slag tab Section 4 for analysis."))
    _build_raw_tab(wb, "Phase_Map", "phase_map", intro=_intro(
        "RAW TC-PYTHON \u2014 Ternary composition phase mapping at 1800 K",
        "20\u00d720 grid for Cu-Fe-O, Cu-Al-O, Cu-Mn-O, Cu-V-O systems. "
        "Shows which phases are stable at each composition."))


def build_results_tabs(wb):
    """Build 3 derived results tabs (green) with intro cards + toxicity."""
    ws_st, n_st = _build_raw_tab(wb, "Screening_Table", "screening",
        tab_color=GREEN_TAB, intro=_intro(
        "BINARY SCREENING TABLE \u2014 One row per candidate oxide",
        "\u0394G_rxn from TC-Python TCOX14 at three temperatures. "
        "Column K = toxicity hazard. RED = HIGH hazard \u2014 "
        "additional safety review required before experimental use."))
    # Toxicity highlighting (column K = 11)
    if n_st > 0:
        _highlight_toxicity(ws_st, col_idx=11,
                            start_row=INTRO_ROWS + 2,
                            end_row=INTRO_ROWS + 1 + n_st)

    ws_ac, n_ac = _build_raw_tab(wb, "Activity_Corrected", "corrected",
        tab_color=GREEN_TAB, intro=_intro(
        "ACTIVITY-CORRECTED \u0394G \u2014 Accounting for dilute Cu in "
        "liquid Fe",
        "Pre-computed at X_Cu=0.003, \u03b3_Cu=8.5. Column M verdict "
        "is a live formula. See Activity_and_Slag tab for interactive "
        "version with editable parameter cells."))
    # Replace verdict_corrected column (col 13 = M) with tiered IF formula
    # referencing dG_corrected_pO2_1atm_kJ (col 8 = H)
    if n_ac > 0:
        verdict_col = 13  # column M
        dg_col_letter = "H"  # dG_corrected_pO2_1atm_kJ
        for r in range(INTRO_ROWS + 2, INTRO_ROWS + 2 + n_ac):
            ws_ac.cell(row=r, column=verdict_col,
                       value=f'=IF({dg_col_letter}{r}<-10,"ROBUST",'
                             f'IF({dg_col_letter}{r}<0,"MARGINAL",'
                             f'IF({dg_col_letter}{r}<15,"UNCERTAIN",'
                             f'"UNFAVORABLE")))')


# =====================================================================
# Tab 1: Dashboard
# =====================================================================
def build_dashboard(wb):
    ws = wb.active
    ws.title = "Dashboard"

    # Wide columns for layout
    ws.column_dimensions["A"].width = 3     # left margin
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 14
    ws.column_dimensions["H"].width = 14
    ws.column_dimensions["I"].width = 14
    ws.column_dimensions["J"].width = 3     # right margin

    CARD_W = 9  # B through J
    COL_START = 2
    COL_END = 10

    title_font = Font(bold=True, size=16, color="2F5496")
    subtitle_font = Font(bold=True, size=12, color="333333")
    body_font = Font(size=10, color="333333")
    small_font = Font(size=9, color="666666")
    section_font = Font(bold=True, size=11, color="2F5496")

    r = 1

    # ── Title block ──
    _write_card_line(ws, r, COL_START, COL_END,
                     "Cu Removal from Recycled Steel — CALPHAD Screening Workbook",
                     font=title_font,
                     fill=PatternFill(start_color="D6E4F0", end_color="D6E4F0",
                                      fill_type="solid"))
    r += 1
    _write_card_line(ws, r, COL_START, COL_END,
                     "MSE 4381 Capstone | Honda R&D Americas | TCOX14 Database | Spring 2026",
                     font=Font(size=10, color="2F5496"),
                     fill=PatternFill(start_color="D6E4F0", end_color="D6E4F0",
                                      fill_type="solid"))
    r += 2

    # ── Project Summary ──
    _write_card_line(ws, r, COL_START, COL_END,
                     "PROJECT SUMMARY", font=section_font)
    r += 1
    summary_lines = [
        "Recycled steel contains 0.25-0.3% Cu from wiring and motors, "
        "causing hot shortness above 0.1%.",
        "This workbook screens 17 ceramic oxides for Cu removal via "
        "ternary compound formation (Cu + MOx + O2 -> CuMOy).",
        "All 16 modeled ternary reactions are thermodynamically favorable "
        "(negative dG). Top candidates: CuFe2O4 (-112 kJ), "
        "Cu3V2O8 (-109 kJ), CuMn2O4 (-64 kJ).",
        "At steelmaking temperatures (1800K), Cu capture occurs via "
        "dissolution into oxide slag, not crystalline compound formation.",
    ]
    for line in summary_lines:
        _write_card_line(ws, r, COL_START, COL_END, line, font=body_font)
        r += 1
    r += 1

    # ── How to Navigate ──
    _write_card_line(ws, r, COL_START, COL_END,
                     "HOW TO USE THIS WORKBOOK", font=section_font)
    r += 1
    _write_card_line(ws, r, COL_START, COL_END,
                     "Tabs are color-coded. Click any formula cell in a "
                     "blue tab to trace it back to source data.",
                     font=body_font)
    r += 2

    # ── Table of Contents ──
    toc_headers = ["Tab", "Purpose", "Key Contents"]
    toc_ncols = len(toc_headers)
    toc_start_col = COL_START

    for c, h in enumerate(toc_headers):
        ws.cell(row=r, column=toc_start_col + c, value=h)
    style_header(ws, r, toc_ncols + toc_start_col - 1)
    # Fix: only style the actual TOC columns
    for c in range(1, toc_start_col):
        ws.cell(row=r, column=c).fill = WHITE_FILL
        ws.cell(row=r, column=c).font = Font(size=10)
    r += 1

    # Color fills for TOC — applied directly to Tab name cell
    blue_indicator = PatternFill(start_color="4472C4", end_color="4472C4",
                                 fill_type="solid")
    green_indicator = PatternFill(start_color="70AD47", end_color="70AD47",
                                  fill_type="solid")
    gray_indicator = PatternFill(start_color="A5A5A5", end_color="A5A5A5",
                                  fill_type="solid")

    toc_entries = [
        ("Dashboard",          blue_indicator,  "Analysis",
         "This tab. Project overview, navigation, key figures."),
        ("Screening_Summary",  blue_indicator,  "Analysis",
         "One row per oxide: binary verdict, ternary verdict, tier rank."),
        ("Mass_Balance",       blue_indicator,  "Analysis",
         "Experiment design: oxide dropdown, particle size sweep, contact time."),
        ("Activity_and_Slag",  blue_indicator,  "Analysis",
         "Dilute Cu penalty, gamma sensitivity, slag basicity effects."),
        ("CuFe2O4_DeepDive",   blue_indicator,  "Analysis",
         "Decompose CuFe2O4 reaction into Cu-capture vs Fe-oxidation."),
        ("Normalization",      blue_indicator,  "Analysis",
         "Full stoichiometry (17 binary + 18 ternary) + 2 worked examples."),
        ("Screening_Table",    green_indicator, "Results",
         "Binary oxide dG at 3 temperatures + physical properties."),
        ("Activity_Corrected", green_indicator, "Results",
         "Activity-corrected dG for all products across temperature."),
        ("Ternary_Verdicts",   gray_indicator,  "Raw/Backing",
         "Ternary product dG + phase data (backing for Screening_Summary)."),
        ("OX_Gibbs",           gray_indicator,  "Raw/Backing",
         "Oxide Gibbs energies (18 oxides, 500-2000K, TCOX14)."),
        ("Ternary_Rxns",       gray_indicator,  "Raw/Backing",
         "Ternary reaction energies (16 systems x 23 temps)."),
        ("dG_Top6",            gray_indicator,  "Raw/Backing",
         "Fine-resolution dG vs T for top 6 candidates (25K steps)."),
        ("CuFe2O4_Raw",        gray_indicator,  "Raw/Backing",
         "CuFe2O4 formation: Cu-capture vs Fe-oxidation components."),
        ("Cu_Activity",        gray_indicator,  "Raw/Backing",
         "Cu activity sweep at 1800K (4 ternary systems)."),
        ("Slag_Effects",       gray_indicator,  "Raw/Backing",
         "Slag basicity effects on a_Cu (2 quaternary systems)."),
        ("Phase_Map",          gray_indicator,  "Raw/Backing",
         "20x20 ternary composition grid phase mapping at 1800K."),
    ]

    toc_first_data = r
    for tab_name, color_fill, purpose, contents in toc_entries:
        tab_cell = ws.cell(row=r, column=toc_start_col, value=tab_name)
        tab_cell.font = Font(size=10, bold=True, color="FFFFFF")
        tab_cell.fill = color_fill
        tab_cell.alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=toc_start_col + 1, value=purpose).font = \
            Font(size=10)
        ws.cell(row=r, column=toc_start_col + 2, value=contents).font = \
            Font(size=9, color="444444")
        r += 1

    toc_last = r - 1
    apply_borders(ws, toc_first_data - 1, toc_last,
                  toc_start_col + toc_ncols - 1)
    r += 1

    # ── Data Provenance ──
    _write_card_line(ws, r, COL_START, COL_END,
                     "DATA PROVENANCE", font=section_font)
    r += 1

    GITHUB_BASE = "https://github.com/dimascad/honda-calphad/blob/main/"

    prov_headers = ["Source Tab", "Source CSV", "Database", "Date",
                    "Generating Script"]
    prov_ncols = len(prov_headers)
    for c, h in enumerate(prov_headers):
        ws.cell(row=r, column=COL_START + c, value=h)
    style_header(ws, r, COL_START + prov_ncols - 1)
    for c in range(1, COL_START):
        ws.cell(row=r, column=c).fill = WHITE_FILL
        ws.cell(row=r, column=c).font = Font(size=10)
    r += 1

    provenance = [
        # (tab, csv, database, date, script_path_in_repo)
        ("OX_Gibbs", "oxide_gibbs_energies.csv", "TCOX14+SSUB3", "Mar 12",
         "simulations/tcpython/extract_oxide_gibbs.py"),
        ("Ternary_Rxns", "ternary_reaction_energies.csv", "TCOX14", "Mar 13",
         "simulations/tcpython/extract_ternary_reactions.py"),
        ("dG_Top6", "dG_vs_T_top6.csv", "TCOX14", "Mar 13",
         "simulations/tcpython/dG_vs_T_top6.py"),
        ("CuFe2O4_Raw", "cufe2o4_alternative_reaction.csv", "TCOX14", "Mar 13",
         "simulations/tcpython/cufe2o4_alternative_reaction.py"),
        ("Cu_Activity", "cu_activity_vs_oxide.csv", "TCOX14", "Mar 13",
         "simulations/tcpython/cu_activity_vs_oxide.py"),
        ("Slag_Effects", "slag_composition_effects.csv", "TCOX14", "Mar 13",
         "simulations/tcpython/slag_composition_effects.py"),
        ("Phase_Map", "ternary_phase_map_1800K.csv", "TCOX14", "Mar 13",
         "simulations/tcpython/ternary_phase_map_1800K.py"),
        ("Activity_Corrected", "activity_corrected_dG.csv", "(derived)", "Mar 13",
         "screening/compute_activity_corrected_dG.py"),
        ("Screening_Table", "screening_table.csv", "(derived)", "Mar 13",
         "screening/build_combined_screening.py"),
        ("Ternary_Verdicts", "ternary_screening_results.csv", "(derived)", "Mar 13",
         "screening/compute_ternary_dG.py"),
    ]
    prov_first = r
    for tab, csv_f, db, date, script in provenance:
        ws.cell(row=r, column=COL_START, value=tab).font = Font(size=9)
        ws.cell(row=r, column=COL_START + 1, value=csv_f).font = Font(size=9)
        ws.cell(row=r, column=COL_START + 2, value=db).font = Font(size=9)
        ws.cell(row=r, column=COL_START + 3, value=date).font = Font(size=9)
        # Script name as clickable GitHub hyperlink
        script_cell = ws.cell(row=r, column=COL_START + 4,
                              value=script.split("/")[-1])
        script_cell.font = Font(size=9, color="0563C1", underline="single")
        script_cell.hyperlink = GITHUB_BASE + script
        r += 1
    prov_last = r - 1
    alt_shading(ws, prov_first, prov_last, COL_START + prov_ncols - 1)
    apply_borders(ws, prov_first - 1, prov_last, COL_START + prov_ncols - 1)
    r += 2

    # ── Key Figures ──
    _write_card_line(ws, r, COL_START, COL_END,
                     "KEY FIGURES", font=section_font)
    r += 1

    # Figure 1: dG vs T
    _write_card_line(ws, r, COL_START, COL_END,
                     "Figure 1. Reaction driving force vs temperature for "
                     "top 6 ternary Cu-capture candidates (25K resolution, "
                     "TCOX14). Data: dG_Top6 tab.",
                     font=Font(size=9, color="444444"))
    r += 1
    rows_consumed = _embed_image(ws, FIGURES["dG_vs_T"], f"B{r}", width_px=620)
    r += rows_consumed + 2

    # Figure 2: Activity correction
    _write_card_line(ws, r, COL_START, COL_END,
                     "Figure 2. Activity-corrected dG comparison: pure "
                     "reference vs dilute Cu (X_Cu=0.003, gamma=8.5). "
                     "Data: Activity_and_Slag tab.",
                     font=Font(size=9, color="444444"))
    r += 1
    rows_consumed = _embed_image(ws, FIGURES["activity"], f"B{r}", width_px=650)
    r += rows_consumed + 2

    # Figure 3: CuFe2O4 decomposition
    _write_card_line(ws, r, COL_START, COL_END,
                     "Figure 3. CuFe2O4 reaction decomposition: total dG "
                     "split into Cu-capture and Fe-oxidation contributions. "
                     "Data: CuFe2O4_DeepDive tab.",
                     font=Font(size=9, color="444444"))
    r += 1
    rows_consumed = _embed_image(ws, FIGURES["cufe2o4"], f"B{r}", width_px=650)
    r += rows_consumed + 2

    # Figure 4: Slag basicity
    _write_card_line(ws, r, COL_START, COL_END,
                     "Figure 4. Effect of slag basicity on Cu activity: "
                     "MnO:SiO2 and Al2O3:SiO2 ratios at 1800K. "
                     "Data: Slag_Effects tab.",
                     font=Font(size=9, color="444444"))
    r += 1
    rows_consumed = _embed_image(ws, FIGURES["slag"], f"B{r}", width_px=620)
    r += rows_consumed + 2

    # Footer
    _write_card_line(ws, r, COL_START, COL_END,
                     "Generated by build_unified_xlsx.py | All plots: "
                     "matplotlib (colorblind-friendly palette per CLAUDE.md)",
                     font=Font(size=8, italic=True, color="999999"))

    freeze_and_color(ws, DASH_TAB, freeze_row=3)
    print(f"  Dashboard: project summary + TOC ({len(toc_entries)} tabs) + "
          f"4 embedded figures")
    return ws


# =====================================================================
# Tab 2: Screening_Summary
# =====================================================================
def build_screening_summary(wb):
    ws = wb.create_sheet("Screening_Summary")

    fields_s, rows_s = load_csv(CSV_FILES["screening"])
    fields_t, rows_t = load_csv(CSV_FILES["ternary_v"])

    if not rows_s:
        ws.cell(row=1, column=1, value="screening_table.csv not found")
        freeze_and_color(ws, BLUE_TAB)
        return ws

    # ── Intro card ──
    _write_card_line(ws, 1, 1, 11,
                     "SCREENING SUMMARY — Binary oxide stability + ternary "
                     "Cu-capture potential for 17 candidate oxides",
                     font=Font(bold=True, size=11, color="2F5496"))
    _write_card_line(ws, 2, 1, 11,
                     "All values reference the green/gray backing tabs. "
                     "Click any formula cell to trace its source.",
                     font=Font(size=9, italic=True, color="666666"))

    HDR = 3
    headers = [
        "Oxide", "Formula", "Density (g/cm\u00b3)", "Melting Point (\u00b0C)",
        "Floats on Steel?",
        "\u0394G_binary @ 1527\u00b0C (kJ/mol O\u2082)",
        "Binary Verdict",
        "Best Ternary Product",
        "\u0394G_ternary @ 1527\u00b0C (kJ)",
        "Ternary Verdict", "Tier",
    ]
    ncols = len(headers)
    for c, h in enumerate(headers, 1):
        ws.cell(row=HDR, column=c, value=h)
    style_header(ws, HDR, ncols)

    # Row mappings
    screening_row_map = {}
    for i, r in enumerate(rows_s):
        screening_row_map[r["Formula"]] = i + 2 + INTRO_ROWS

    # Best ternary product per oxide
    oxide_to_ternary = {}
    if rows_t:
        oxide_products = {}
        for i, r in enumerate(rows_t):
            oxide = r["Oxide"]
            dg_str = r.get("dG_rxn @ 1527\u00b0C (kJ)", "").strip()
            if dg_str and dg_str != "No data":
                try:
                    dg = float(dg_str)
                except ValueError:
                    continue
                if oxide not in oxide_products:
                    oxide_products[oxide] = []
                oxide_products[oxide].append((dg, r["Product"], i + 2 + INTRO_ROWS))
        for oxide, prods in oxide_products.items():
            prods.sort()
            oxide_to_ternary[oxide] = (prods[0][1], prods[0][2])

    data_start = HDR + 1
    for i, r in enumerate(rows_s):
        row = data_start + i
        formula = r["Formula"]
        st_row = screening_row_map.get(formula, 0)

        ws.cell(row=row, column=1, value=f"=Screening_Table!A{st_row}")
        ws.cell(row=row, column=2, value=f"=Screening_Table!B{st_row}")
        ws.cell(row=row, column=3, value=f"=Screening_Table!C{st_row}")
        ws.cell(row=row, column=4, value=f"=Screening_Table!D{st_row}")
        ws.cell(row=row, column=5, value=f"=Screening_Table!E{st_row}")
        ws.cell(row=row, column=6, value=f"=Screening_Table!I{st_row}")
        ws.cell(row=row, column=7,
                value=f'=IF(ISNUMBER(F{row}),IF(F{row}<0,'
                      f'"Cu REDUCES","STABLE"),"N/A")')

        if formula in oxide_to_ternary:
            prod, tv_row = oxide_to_ternary[formula]
            ws.cell(row=row, column=8,
                    value=f"=Ternary_Verdicts!A{tv_row}")
            ws.cell(row=row, column=9,
                    value=f"=Ternary_Verdicts!G{tv_row}")
            ws.cell(row=row, column=10,
                    value=f'=IF(ISNUMBER(I{row}),IF(I{row}<0,'
                          f'"Cu CAPTURED","NO CAPTURE"),"N/A")')
            ws.cell(row=row, column=11,
                    value=f'=IF(ISNUMBER(I{row}),'
                          f'IF(I{row}<-50,"STRONG",'
                          f'IF(I{row}<-25,"MODERATE","WEAK")),"N/A")')
        else:
            ws.cell(row=row, column=8, value="No data")
            ws.cell(row=row, column=9, value="N/A")
            ws.cell(row=row, column=10, value="N/A")
            ws.cell(row=row, column=11, value="N/A")

    last_data = data_start + len(rows_s) - 1

    alt_shading(ws, data_start, last_data, ncols)
    center_all(ws, HDR, last_data, ncols)
    apply_borders(ws, HDR, last_data, ncols)

    # ── TOP CANDIDATES sub-table ──
    gap = last_data + 2
    _write_card_line(ws, gap, 1, ncols,
                     "TOP CANDIDATES", font=Font(bold=True, size=12))

    top_hdr = gap + 1
    top_headers = [
        "Product", "Parent Oxide", "Reaction",
        "\u0394G @ 1527\u00b0C (kJ)", "Named Phase",
        "Mechanism", "Tier",
    ]
    top_ncols = len(top_headers)
    for c, h in enumerate(top_headers, 1):
        ws.cell(row=top_hdr, column=c, value=h)
    style_header(ws, top_hdr, top_ncols)

    candidates = [
        ("CuFe\u2082O\u2084", "FeO",
         "Cu + 2FeO + O\u2082 \u2192 CuFe\u2082O\u2084",
         "SPINEL#1 (to 1700K)", "Spinel trapping", "STRONG"),
        ("Cu\u2083V\u2082O\u2088", "V\u2082O\u2085",
         "3Cu + V\u2082O\u2085 + 1.5O\u2082 \u2192 Cu\u2083V\u2082O\u2088",
         "none (liquid)", "Vanadate slag dissolution", "STRONG"),
        ("CuMn\u2082O\u2084", "MnO",
         "Cu + 2MnO + O\u2082 \u2192 CuMn\u2082O\u2084",
         "SPINEL#1 (to 1700K)", "Spinel trapping", "STRONG"),
        ("Cu\u2082SiO\u2084", "SiO\u2082",
         "2Cu + SiO\u2082 + O\u2082 \u2192 Cu\u2082SiO\u2084",
         "none (liquid)", "Silicate slag dissolution", "STRONG"),
    ]

    prod_row_map = {}
    if rows_t:
        for idx, r in enumerate(rows_t):
            prod_row_map[r["Product"]] = idx + 2 + INTRO_ROWS

    prod_csv_names = {
        "CuFe\u2082O\u2084": "CuFe2O4",
        "Cu\u2083V\u2082O\u2088": "Cu3V2O8",
        "CuMn\u2082O\u2084": "CuMn2O4",
        "Cu\u2082SiO\u2084": "Cu2SiO4",
    }

    for j, (prod, oxide, rxn, phase, mech, tier) in enumerate(candidates):
        row = top_hdr + 1 + j
        ws.cell(row=row, column=1, value=prod)
        ws.cell(row=row, column=2, value=oxide)
        ws.cell(row=row, column=3, value=rxn)
        csv_name = prod_csv_names.get(prod, "")
        if csv_name in prod_row_map:
            tv_r = prod_row_map[csv_name]
            ws.cell(row=row, column=4,
                    value=f"=Ternary_Verdicts!G{tv_r}")
        else:
            ws.cell(row=row, column=4, value="N/A")
        ws.cell(row=row, column=5, value=phase)
        ws.cell(row=row, column=6, value=mech)
        ws.cell(row=row, column=7, value=tier)

    top_last = top_hdr + len(candidates)
    alt_shading(ws, top_hdr + 1, top_last, top_ncols)
    center_all(ws, top_hdr, top_last, top_ncols)
    apply_borders(ws, top_hdr, top_last, top_ncols)

    auto_width(ws, ncols, max_w=35)

    # ── Oxide highlight selector ──
    # Place dropdown in row 1, col M (outside merged intro area)
    HL_COL = 13  # col M
    ws.cell(row=1, column=HL_COL,
            value="Highlight:").font = Font(size=9, bold=True,
                                            color="2F5496")
    ws.cell(row=1, column=HL_COL).alignment = Alignment(horizontal="right")
    highlight_cell = ws.cell(row=1, column=HL_COL + 1, value="(none)")
    highlight_cell.fill = DROPDOWN_FILL
    highlight_cell.font = DROPDOWN_FONT
    highlight_cell.border = DROPDOWN_BORDER
    ws.column_dimensions[get_column_letter(HL_COL)].width = 10
    ws.column_dimensions[get_column_letter(HL_COL + 1)].width = 14

    # Build formula list from the oxide names in the data
    oxide_formulas = [r["Formula"] for r in rows_s]
    dv_list = "(none)," + ",".join(oxide_formulas[:25])  # DataValidation limit
    highlight_dv = DataValidation(
        type="list",
        formula1=f'"{dv_list}"',
        allow_blank=True,
    )
    highlight_dv.prompt = "Select an oxide to highlight its row in yellow."
    highlight_dv.promptTitle = "Highlight Oxide"
    ws.add_data_validation(highlight_dv)
    highlight_dv.add(highlight_cell)

    # Conditional formatting: highlight entire row when col B matches $N$1
    from openpyxl.formatting.rule import FormulaRule
    hl_ref = f"${get_column_letter(HL_COL + 1)}$1"
    highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00",
                                  fill_type="solid")
    highlight_range = f"A{data_start}:K{last_data}"
    ws.conditional_formatting.add(
        highlight_range,
        FormulaRule(formula=[f'=$B{data_start}={hl_ref}'],
                    fill=highlight_fill))

    # Conditional formatting for verdict columns
    for col_l in ["G", "J"]:
        cell_range = f"{col_l}{data_start}:{col_l}{last_data}"
        ws.conditional_formatting.add(
            cell_range,
            CellIsRule(operator="equal", formula=['"Cu REDUCES"'],
                       fill=RED_FILL))
        ws.conditional_formatting.add(
            cell_range,
            CellIsRule(operator="equal", formula=['"Cu CAPTURED"'],
                       fill=GREEN_FILL))

    freeze_and_color(ws, BLUE_TAB, freeze_row=4)
    print(f"  Screening_Summary: {len(rows_s)} oxides + {len(candidates)} "
          f"top candidates")
    return ws


# =====================================================================
# Tab 3: CuFe2O4_DeepDive
# =====================================================================
def build_cufe2o4_deepdive(wb):
    ws = wb.create_sheet("CuFe2O4_DeepDive")

    fields_raw, rows_raw = load_csv(CSV_FILES["cufe2o4"])
    if not rows_raw:
        ws.cell(row=1, column=1, value="CuFe2O4 data not available")
        freeze_and_color(ws, BLUE_TAB)
        return ws

    nrows = len(rows_raw)

    # ── Intro card ──
    _write_card_line(ws, 1, 1, 7,
                     "CuFe2O4 DEEP DIVE — Decomposing the strongest candidate "
                     "into Cu-capture vs Fe-oxidation",
                     font=Font(bold=True, size=11, color="2F5496"))
    _write_card_line(ws, 2, 1, 7,
                     "The total CuFe2O4 formation reaction (Cu + 2FeO + O2) "
                     "is split into two contributions. Columns E-G are "
                     "formulas referencing the CuFe2O4_Raw tab.",
                     font=Font(size=9, italic=True, color="666666"))

    HDR = 3
    headers = [
        "T (K)", "T (\u00b0C)",
        "dG_original (kJ)", "dG_alternative (kJ)",
        "dG_Fe_oxidation (kJ)", "Cu-capture (%)", "Fe-oxidation (%)",
    ]
    ncols = len(headers)
    for c, h in enumerate(headers, 1):
        ws.cell(row=HDR, column=c, value=h)
    style_header(ws, HDR, ncols)

    data_start = HDR + 1
    for i in range(nrows):
        row = data_start + i
        raw_row = i + 2 + INTRO_ROWS

        ws.cell(row=row, column=1, value=f"=CuFe2O4_Raw!A{raw_row}")
        ws.cell(row=row, column=2, value=f"=A{row}-273.15")
        fmt_number(ws, row, 2, "0.00")
        ws.cell(row=row, column=3, value=f"=CuFe2O4_Raw!L{raw_row}")
        fmt_number(ws, row, 3, "+0.00;-0.00;0.00")
        ws.cell(row=row, column=4, value=f"=CuFe2O4_Raw!J{raw_row}")
        fmt_number(ws, row, 4, "+0.00;-0.00;0.00")
        ws.cell(row=row, column=5, value=f"=C{row}-D{row}")
        fmt_number(ws, row, 5, "+0.00;-0.00;0.00")
        ws.cell(row=row, column=6,
                value=f"=IF(C{row}<>0,D{row}/C{row}*100,0)")
        fmt_number(ws, row, 6, "0.0")
        ws.cell(row=row, column=7, value=f"=100-F{row}")
        fmt_number(ws, row, 7, "0.0")

    last_data = data_start + nrows - 1

    # AVERAGE row
    avg_row = last_data + 1
    ws.cell(row=avg_row, column=1, value="AVERAGE")
    ws.cell(row=avg_row, column=1).font = Font(bold=True, size=10)
    for c in range(3, ncols + 1):
        col_l = get_column_letter(c)
        ws.cell(row=avg_row, column=c,
                value=f"=AVERAGE({col_l}{data_start}:{col_l}{last_data})")
        fmt_number(ws, avg_row, c,
                   "+0.00;-0.00;0.00" if c <= 5 else "0.0")
    apply_row_fill(ws, avg_row, ncols, GRAY_FILL)

    alt_shading(ws, data_start, last_data, ncols)
    center_all(ws, HDR, avg_row, ncols)
    apply_borders(ws, HDR, avg_row, ncols)
    auto_width(ws, ncols)

    # ── Excel chart: dG decomposition vs temperature ──
    chart = LineChart()
    chart.title = "CuFe\u2082O\u2084 Reaction Decomposition"
    chart.x_axis.title = "Temperature (\u00b0C)"
    chart.y_axis.title = "\u0394G (kJ/mol)"
    chart.width = 22
    chart.height = 13
    chart.style = 10

    # X-axis: T(°C) in col B
    x_data = Reference(ws, min_col=2, min_row=data_start,
                       max_row=last_data)

    # Series: dG_original (C), dG_alternative (D), dG_Fe_oxidation (E)
    series_info = [
        (3, "dG total (original)"),
        (4, "dG Cu-capture (alternative)"),
        (5, "dG Fe-oxidation"),
    ]
    colors_chart = ["2F5496", "AA3377", "70AD47"]

    for (col, title), color in zip(series_info, colors_chart):
        data_ref = Reference(ws, min_col=col, min_row=HDR,
                             max_row=last_data)
        chart.add_data(data_ref, titles_from_data=True)

    chart.set_categories(x_data)

    # Style the series — use CLAUDE.md colorblind palette for 3 series
    deepdive_styles = [
        ("0077BB", "solid"),   # Blue — total
        ("AA3377", "dash"),    # Purple — Cu-capture
        ("70AD47", "dashDot"), # Green — Fe-oxidation
    ]
    for idx, (color, dash) in enumerate(deepdive_styles):
        s = chart.series[idx]
        s.graphicalProperties.line.width = 22000  # ~1.7pt in EMU
        s.graphicalProperties.line.solidFill = color
        s.graphicalProperties.line.dashStyle = dash

    # Axis visibility and tick labels
    chart.x_axis.delete = False
    chart.y_axis.delete = False
    chart.x_axis.tickLblPos = "low"
    chart.y_axis.tickLblPos = "low"
    chart.y_axis.crossesAt = 0
    # Gridlines on both axes (light gray)
    grid_sp = GraphicalProperties(ln=LineProperties(solidFill="D0D0D0", w=6350))
    chart.y_axis.majorGridlines = ChartLines(spPr=grid_sp)
    chart.x_axis.majorGridlines = ChartLines(spPr=grid_sp)

    ws.add_chart(chart, f"A{avg_row + 2}")

    freeze_and_color(ws, BLUE_TAB, freeze_row=4)

    print(f"  CuFe2O4_DeepDive: {nrows} rows, formula cols E-G + AVERAGE "
          f"+ Excel chart")
    return ws


# =====================================================================
# Tab 4: Activity_and_Slag
# =====================================================================
def build_activity_and_slag(wb):
    ws = wb.create_sheet("Activity_and_Slag")

    note_font = Font(italic=True, size=9, color="666666")

    # ── Intro card ──
    _write_card_line(ws, 1, 1, 8,
                     "ACTIVITY AND SLAG EFFECTS — What happens under real "
                     "steelmaking conditions?",
                     font=Font(bold=True, size=11, color="2F5496"))
    _write_card_line(ws, 2, 1, 8,
                     "Section 1: Dilute Cu penalty (editable X_Cu and "
                     "gamma_Cu). Section 2: Gamma sensitivity sweep. "
                     "Section 3: Slag basicity effects from Slag_Effects tab.",
                     font=Font(size=9, italic=True, color="666666"))

    # ── Section 1: Parameters ──
    param_font = Font(bold=True, size=10)

    ws.cell(row=4, column=1, value="Parameter").font = param_font
    ws.cell(row=4, column=2, value="Symbol").font = param_font
    ws.cell(row=4, column=3, value="Value").font = param_font
    ws.cell(row=4, column=4, value="Unit / Note").font = param_font
    style_header(ws, 4, 4)

    ws.cell(row=5, column=1, value="Cu mole fraction in steel")
    ws.cell(row=5, column=2, value="X_Cu")
    xcu_cell = ws.cell(row=5, column=3, value=0.003)
    xcu_cell.fill = DROPDOWN_FILL
    xcu_cell.font = DROPDOWN_FONT
    xcu_cell.border = DROPDOWN_BORDER
    fmt_number(ws, 5, 3, "0.0000")
    ws.cell(row=5, column=4, value="~0.3 wt% Cu (EDITABLE)")

    ws.cell(row=6, column=1, value="Raoultian activity coefficient")
    ws.cell(row=6, column=2, value="\u03b3_Cu")
    gamma_cell = ws.cell(row=6, column=3, value=8.5)
    gamma_cell.fill = DROPDOWN_FILL
    gamma_cell.font = DROPDOWN_FONT
    gamma_cell.border = DROPDOWN_BORDER
    fmt_number(ws, 6, 3, "0.0")
    ws.cell(row=6, column=4, value="Literature range: 5-13 (EDITABLE)")

    ws.cell(row=7, column=1, value="Cu activity in liquid Fe")
    ws.cell(row=7, column=2, value="a_Cu")
    ws.cell(row=7, column=3, value="=C5*C6")
    fmt_number(ws, 7, 3, "0.0000")
    ws.cell(row=7, column=4, value="X_Cu \u00d7 \u03b3_Cu (formula in C7)")

    ws.cell(row=8, column=1, value="Temperature")
    ws.cell(row=8, column=2, value="T")
    t_cell = ws.cell(row=8, column=3, value=1800)
    t_cell.fill = DROPDOWN_FILL
    t_cell.font = DROPDOWN_FONT
    t_cell.border = DROPDOWN_BORDER
    ws.cell(row=8, column=4, value="K \u25bc SELECT FROM DROPDOWN")

    t_dv = DataValidation(
        type="list",
        formula1='"1400,1500,1600,1700,1800,1900,2000"',
        allow_blank=False,
    )
    t_dv.prompt = ("Select temperature. The penalty term -RT ln(a_Cu) "
                   "updates automatically.")
    t_dv.promptTitle = "Temperature"
    ws.add_data_validation(t_dv)
    t_dv.add(t_cell)

    ws.cell(row=9, column=1, value="Gas constant")
    ws.cell(row=9, column=2, value="R")
    ws.cell(row=9, column=3, value=8.314)
    fmt_number(ws, 9, 3, "0.000")
    ws.cell(row=9, column=4, value="J/(mol\u00b7K)")

    ws.cell(row=10, column=1, value="Penalty per Cu atom consumed")
    ws.cell(row=10, column=2, value="-RT\u00b7ln(a_Cu)")
    ws.cell(row=10, column=3, value="=-C9*C8*LN(C7)/1000")
    fmt_number(ws, 10, 3, "+0.0;-0.0")
    ws.cell(row=10, column=4, value="kJ/mol (positive = less favorable)")

    alt_shading(ws, 5, 10, 4)
    apply_borders(ws, 4, 10, 4)

    # ── Section 2: Product correction table ──
    TABLE_START = 12
    _write_card_line(ws, TABLE_START, 1, 8,
                     "ACTIVITY-CORRECTED \u0394G AT 1800 K",
                     font=Font(bold=True, size=11))

    hdr = TABLE_START + 1
    col_headers = [
        "Product", "Parent Oxide", "n_Cu",
        "\u0394G\u00b0 (kJ)", "Cu Penalty (kJ)",
        "\u0394G_eff (kJ)", "Verdict", "\u0394G per Cu (kJ)",
    ]
    prod_ncols = len(col_headers)
    for c, h in enumerate(col_headers, 1):
        ws.cell(row=hdr, column=c, value=h)
    style_header(ws, hdr, prod_ncols)

    _, corr_rows = load_csv(CSV_FILES["corrected"])
    rows_1800 = [r for r in corr_rows if abs(float(r["T_K"]) - 1800) < 1]
    rows_1800.sort(key=lambda r: float(r["dG_corrected_pO2_1atm_kJ"]))

    # Build product -> Ternary_Rxns row map at T=1800K
    _, tern_rows = load_csv(CSV_FILES["ternary_rxns"])
    prod_tern_row = {}
    for i, tr in enumerate(tern_rows):
        if abs(float(tr["T_K"]) - 1800) < 1:
            prod_tern_row[tr["product"]] = i + 2 + INTRO_ROWS

    data_start = hdr + 1
    for i, r in enumerate(rows_1800):
        row = data_start + i
        n_cu = int(r["n_Cu"])
        product = r["product"]

        ws.cell(row=row, column=1, value=product)
        ws.cell(row=row, column=2, value=r.get("parent_oxide", ""))
        ws.cell(row=row, column=3, value=n_cu)
        # dG_pure: reference Ternary_Rxns col N (dG_rxn_system_kJ)
        tr_row = prod_tern_row.get(product)
        if tr_row:
            ws.cell(row=row, column=4,
                    value=f"=Ternary_Rxns!N{tr_row}")
        else:
            ws.cell(row=row, column=4, value=float(r["dG_pure_kJ"]))
        fmt_number(ws, row, 4, "+0.0;-0.0;0.0")
        ws.cell(row=row, column=5, value=f"=C{row}*$C$10")
        fmt_number(ws, row, 5, "+0.0;-0.0;0.0")
        ws.cell(row=row, column=6, value=f"=D{row}+E{row}")
        fmt_number(ws, row, 6, "+0.0;-0.0;0.0")
        ws.cell(row=row, column=7,
                value=f'=IF(F{row}<-10,"ROBUST",'
                      f'IF(F{row}<0,"MARGINAL",'
                      f'IF(F{row}<15,"UNCERTAIN","UNFAVORABLE")))')
        ws.cell(row=row, column=8, value=f"=D{row}/C{row}")
        fmt_number(ws, row, 8, "+0.0;-0.0;0.0")

    last_data = data_start + len(rows_1800) - 1
    alt_shading(ws, data_start, last_data, prod_ncols)
    center_all(ws, hdr, last_data, prod_ncols)
    apply_borders(ws, hdr, last_data, prod_ncols)

    main_range = f"F{data_start}:F{last_data}"
    ws.conditional_formatting.add(
        main_range,
        CellIsRule(operator="lessThan", formula=["0"], fill=GREEN_FILL))
    ws.conditional_formatting.add(
        main_range,
        CellIsRule(operator="greaterThanOrEqual", formula=["0"],
                   fill=RED_FILL))

    # ── Section 3: Gamma sensitivity sweep ──
    SENS_START = last_data + 2
    _write_card_line(ws, SENS_START, 1, 8,
                     "\u03b3_Cu SENSITIVITY ANALYSIS",
                     font=Font(bold=True, size=11))
    ws.cell(row=SENS_START + 1, column=1,
            value="Change \u03b3_Cu in cell C6 above to see all "
                  "corrections update automatically.").font = note_font
    ws.merge_cells(f"A{SENS_START+1}:H{SENS_START+1}")

    sens_hdr = SENS_START + 2
    sens_cols = ["\u03b3_Cu", "a_Cu", "Penalty/Cu (kJ)"]
    product_names = [r["product"] for r in rows_1800]
    for p in product_names:
        sens_cols.append(f"\u0394G_eff {p}")
    sens_ncols = len(sens_cols)

    for c, h in enumerate(sens_cols, 1):
        ws.cell(row=sens_hdr, column=c, value=h)
    style_header(ws, sens_hdr, sens_ncols)

    gamma_values = [3, 4, 5, 6, 7, 8, 8.5, 9, 10, 11, 12, 13, 15]

    for j, gv in enumerate(gamma_values):
        row = sens_hdr + 1 + j
        ws.cell(row=row, column=1, value=gv)
        fmt_number(ws, row, 1, "0.0")
        ws.cell(row=row, column=2, value=f"=A{row}*$C$5")
        fmt_number(ws, row, 2, "0.0000")
        ws.cell(row=row, column=3, value=f"=-$C$9*$C$8*LN(B{row})/1000")
        fmt_number(ws, row, 3, "+0.0;-0.0")

        for k, r_data in enumerate(rows_1800):
            col = 4 + k
            n_cu = int(r_data["n_Cu"])
            # Reference Section 2 dG_pure cell (col D) and n_Cu cell (col C)
            sec2_row = data_start + k
            ws.cell(row=row, column=col,
                    value=f"=$D${sec2_row}+$C${sec2_row}*C{row}")
            fmt_number(ws, row, col, "+0.0;-0.0;0.0")

    sens_last = sens_hdr + len(gamma_values)
    alt_shading(ws, sens_hdr + 1, sens_last, sens_ncols)
    center_all(ws, sens_hdr, sens_last, sens_ncols)
    apply_borders(ws, sens_hdr, sens_last, sens_ncols)

    for k in range(len(rows_1800)):
        col_letter = get_column_letter(4 + k)
        cell_range = f"{col_letter}{sens_hdr+1}:{col_letter}{sens_last}"
        ws.conditional_formatting.add(
            cell_range,
            CellIsRule(operator="lessThan", formula=["0"], fill=GREEN_FILL))
        ws.conditional_formatting.add(
            cell_range,
            CellIsRule(operator="greaterThanOrEqual", formula=["0"],
                       fill=RED_FILL))

    # ── Excel chart: gamma sensitivity ──
    sens_chart = LineChart()
    sens_chart.title = (
        "\u0394G_eff vs \u03b3_Cu at 1800 K "
        "(X_Cu = 0.003)")
    sens_chart.x_axis.title = "\u03b3_Cu (Raoultian activity coefficient)"
    sens_chart.y_axis.title = "\u0394G_eff (kJ/mol)"
    sens_chart.width = 24
    sens_chart.height = 14
    sens_chart.style = 10

    # X-axis: gamma values in col A
    x_cats = Reference(ws, min_col=1, min_row=sens_hdr + 1,
                       max_row=sens_last)

    # Series: cols D through D+len(products)-1 (dG_eff per product)
    n_products = len(rows_1800)
    # Product order in sensitivity table (matches Section 2 sort order)
    product_order_sens = [r["product"] for r in rows_1800]

    for k in range(n_products):
        col = 4 + k
        data_ref = Reference(ws, min_col=col, min_row=sens_hdr,
                             max_row=sens_last)
        sens_chart.add_data(data_ref, titles_from_data=True)

    sens_chart.set_categories(x_cats)

    # Style series — match matplotlib colors and dash styles
    for idx in range(n_products):
        s = sens_chart.series[idx]
        product = product_order_sens[idx]
        color, dash = PRODUCT_CHART_STYLE.get(product, ("333333", "solid"))
        s.graphicalProperties.line.width = 22000
        s.graphicalProperties.line.solidFill = color
        s.graphicalProperties.line.dashStyle = dash

    # Axis visibility and tick labels
    sens_chart.x_axis.delete = False
    sens_chart.y_axis.delete = False
    sens_chart.x_axis.tickLblPos = "low"
    sens_chart.y_axis.tickLblPos = "low"
    sens_chart.y_axis.crossesAt = 0
    # Gridlines on both axes (light gray, matching DeepDive chart)
    grid_sp = GraphicalProperties(ln=LineProperties(solidFill="D0D0D0", w=6350))
    sens_chart.y_axis.majorGridlines = ChartLines(spPr=grid_sp)
    sens_chart.x_axis.majorGridlines = ChartLines(spPr=grid_sp)

    chart_anchor_row = sens_last + 2
    ws.add_chart(sens_chart, f"A{chart_anchor_row}")
    CHART_ROWS = 28  # rows consumed by chart (height=14 ~ 28 rows)

    # ── Section 4: Slag Effects ──
    SLAG_START = chart_anchor_row + CHART_ROWS
    _write_card_line(ws, SLAG_START, 1, 5,
                     "SLAG COMPOSITION EFFECTS ON Cu ACTIVITY",
                     font=Font(bold=True, size=11))

    slag_hdr = SLAG_START + 1
    slag_headers = [
        "System", "Ratio Label", "Ratio Value",
        "a_Cu", "% Change from Baseline",
    ]
    slag_ncols = len(slag_headers)
    for c, h in enumerate(slag_headers, 1):
        ws.cell(row=slag_hdr, column=c, value=h)
    style_header(ws, slag_hdr, slag_ncols)

    _, slag_rows = load_csv(CSV_FILES["slag"])
    nslag = len(slag_rows)

    system_baseline_row = {}
    for i in range(nslag):
        row = slag_hdr + 1 + i
        raw_row = i + 2 + INTRO_ROWS

        ws.cell(row=row, column=1, value=f"=Slag_Effects!A{raw_row}")
        ws.cell(row=row, column=2, value=f"=Slag_Effects!B{raw_row}")
        ws.cell(row=row, column=3, value=f"=Slag_Effects!C{raw_row}")
        fmt_number(ws, row, 3, "0.0")
        ws.cell(row=row, column=4, value=f"=Slag_Effects!I{raw_row}")
        fmt_number(ws, row, 4, "0.0000E+00")

        sys = slag_rows[i]["system"]
        if sys not in system_baseline_row:
            system_baseline_row[sys] = row
            ws.cell(row=row, column=5, value=0.0)
        else:
            base = system_baseline_row[sys]
            ws.cell(row=row, column=5,
                    value=f"=(D{row}-D${base})/D${base}*100")
        fmt_number(ws, row, 5, "+0.00;-0.00;0.00")

    slag_last = slag_hdr + nslag
    group_shading(ws, slag_hdr + 1, slag_last, slag_ncols, group_col=1)
    center_all(ws, slag_hdr, slag_last, slag_ncols)
    apply_borders(ws, slag_hdr, slag_last, slag_ncols)

    auto_width(ws, max(prod_ncols, sens_ncols, slag_ncols), max_w=30)
    ws.column_dimensions["A"].width = 32
    freeze_and_color(ws, BLUE_TAB, freeze_row=5)

    print(f"  Activity_and_Slag: {len(rows_1800)} products, "
          f"{len(gamma_values)} gamma points, {nslag} slag rows")
    return ws


# =====================================================================
# Tab 5: Normalization (full stoichiometry + worked examples)
# =====================================================================
def build_mass_balance(wb):
    """Tab: Experiment design — oxide mass, particles, contact time.

    Features:
      - Editable parameter cells (steel mass, Cu%, excess factor, radius)
      - Oxide comparison table with formula-computed stoichiometry
      - Oxide DROPDOWN selector — particle sweep updates via INDEX/MATCH
      - Particle size sweep for whichever oxide is selected
    """
    ws = wb.create_sheet("Mass_Balance")

    note_font = Font(italic=True, size=9, color="666666")
    param_font = Font(bold=True, size=10)
    section_font = Font(bold=True, size=11, color="2F5496")
    edit_fill = PatternFill(start_color="DAEEF3", end_color="DAEEF3",
                            fill_type="solid")

    # ── Intro card ──
    _write_card_line(ws, 1, 1, 9,
                     "EXPERIMENT DESIGN \u2014 How much oxide, what particle "
                     "size, how long?",
                     font=section_font)
    _write_card_line(ws, 2, 1, 9,
                     "Gold cells are editable. All results update "
                     "automatically. Use the oxide dropdown (C15) to "
                     "switch the particle sweep.",
                     font=note_font)

    # ── Section 1: Editable Parameters (rows 4-15) ──
    ws.cell(row=4, column=1, value="Parameter").font = param_font
    ws.cell(row=4, column=2, value="Symbol").font = param_font
    ws.cell(row=4, column=3, value="Value").font = param_font
    ws.cell(row=4, column=4, value="Unit / Note").font = param_font
    style_header(ws, 4, 4)

    params = [
        # (row, label, symbol, value, fmt, unit, editable)
        (5,  "Steel mass",           "m_steel",    0.5,      "0.00",     "kg (EDITABLE)",               True),
        (6,  "Initial Cu",           "Cu_init",    0.30,     "0.00",     "wt% (EDITABLE)",              True),
        (7,  "Target Cu",            "Cu_target",  0.10,     "0.00",     "wt% (EDITABLE)",              True),
        (8,  "Excess factor",        "k_excess",   3.0,      "0.0",      "\u00d7 stoichiometric (EDITABLE)", True),
        (9,  "Particle radius",      "R",          100,      "0",        "\u03bcm (EDITABLE)",          True),
        (10, "D_Cu in liquid Fe",    "D_Cu",       9.63e-10, "0.00E+00", "m\u00b2/s at 1800K (DICTRA)", False),
        (11, "Liquid steel density", "\u03c1_steel", 7000,   "0",        "kg/m\u00b3",                  False),
    ]
    for row, label, symbol, value, fmt, unit, editable in params:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=symbol)
        c = ws.cell(row=row, column=3, value=value)
        c.number_format = fmt
        if editable:
            c.fill = DROPDOWN_FILL
            c.font = DROPDOWN_FONT
            c.border = DROPDOWN_BORDER
        ws.cell(row=row, column=4, value=unit)

    # Derived parameters (formulas)
    ws.cell(row=12, column=1, value="Cu to remove")
    ws.cell(row=12, column=2, value="\u0394Cu")
    ws.cell(row=12, column=3, value="=C5*1000*(C6-C7)/100")
    fmt_number(ws, 12, 3, "0.00")
    ws.cell(row=12, column=4,
            value="g  =  m_steel \u00d7 1000 \u00d7 (Cu_init \u2212 Cu_target) / 100")

    ws.cell(row=13, column=1, value="Cu moles to remove")
    ws.cell(row=13, column=2, value="n_Cu")
    ws.cell(row=13, column=3, value="=C12/63.546")
    fmt_number(ws, 13, 3, "0.0000")
    ws.cell(row=13, column=4, value="mol  =  \u0394Cu / MW_Cu (63.546 g/mol)")

    ws.cell(row=14, column=1, value="Diffusion time")
    ws.cell(row=14, column=2, value="t_diff")
    ws.cell(row=14, column=3, value="=(C9*1E-6)^2/(2*C10)")
    fmt_number(ws, 14, 3, "0.0")
    ws.cell(row=14, column=4, value="s  =  R\u00b2 / (2 D_Cu)")

    # ── Oxide selector dropdown (row 15) ──
    ws.cell(row=15, column=1, value="Selected oxide")
    ws.cell(row=15, column=2, value="oxide")
    # Unicode oxide names for the dropdown list and default value
    oxide_names = [
        "Fe\u2082O\u2083", "V\u2082O\u2085", "MnO",
        "SiO\u2082", "Al\u2082O\u2083",
    ]
    c15 = ws.cell(row=15, column=3, value=oxide_names[0])  # default Fe2O3
    c15.fill = DROPDOWN_FILL
    c15.font = DROPDOWN_FONT
    c15.border = DROPDOWN_BORDER
    ws.cell(row=15, column=4,
            value="\u25bc SELECT FROM DROPDOWN \u2192 particle sweep updates")

    dv = DataValidation(
        type="list",
        formula1='"' + ",".join(oxide_names) + '"',
        allow_blank=False,
    )
    dv.error = "Pick one of the 5 oxides."
    dv.errorTitle = "Invalid oxide"
    dv.prompt = "Select an oxide to analyze in the particle sweep below."
    dv.promptTitle = "Oxide Selection"
    ws.add_data_validation(dv)
    dv.add(c15)

    # ── Section 2: Oxide Comparison Table (rows 18+) ──
    OX_TABLE_ROW = 18  # header row
    ws.cell(row=OX_TABLE_ROW - 1, column=1,
            value="OXIDE COMPARISON TABLE").font = section_font

    headers = ["Oxide", "Product", "\u0394G(1800K)", "MW_oxide",
               "\u03c1 (kg/m\u00b3)", "Cu/mol oxide",
               "Stoich (g)", "Recommended (g)", "wt% of steel"]
    for ci, h in enumerate(headers, 1):
        ws.cell(row=OX_TABLE_ROW, column=ci, value=h)
    style_header(ws, OX_TABLE_ROW, len(headers))

    oxides = [
        # (name, product, dG, MW, rho, cu_per_mol)
        ("Fe\u2082O\u2083", "CuFe\u2082O\u2084", -111.9, 159.69, 5240, 1),
        ("V\u2082O\u2085",  "Cu\u2083V\u2082O\u2088", -109.2, 181.88, 3357, 3),
        ("MnO",             "CuMn\u2082O\u2084",  -63.5,  70.94, 5430, 0.5),
        ("SiO\u2082",       "Cu\u2082SiO\u2084",  -50.2,  60.08, 2650, 2),
        ("Al\u2082O\u2083", "CuAl\u2082O\u2084",  -35.7, 101.96, 3950, 1),
    ]

    first_data_row = OX_TABLE_ROW + 1
    for i, (name, product, dG, mw, rho, cu_per_mol) in enumerate(oxides):
        row = first_data_row + i
        ws.cell(row=row, column=1, value=name)      # A: oxide name (lookup key)
        ws.cell(row=row, column=2, value=product)    # B: product
        ws.cell(row=row, column=3, value=dG)         # C: dG
        fmt_number(ws, row, 3, "0.0")
        ws.cell(row=row, column=4, value=mw)         # D: MW
        fmt_number(ws, row, 4, "0.00")
        ws.cell(row=row, column=5, value=rho)        # E: density
        ws.cell(row=row, column=6, value=cu_per_mol) # F: Cu per mol
        # G: stoich = n_Cu / cu_per_mol * MW_oxide
        ws.cell(row=row, column=7, value=f"=$C$13/F{row}*D{row}")
        fmt_number(ws, row, 7, "0.00")
        # H: recommended = stoich * excess
        ws.cell(row=row, column=8, value=f"=G{row}*$C$8")
        fmt_number(ws, row, 8, "0.00")
        # I: wt% of steel
        ws.cell(row=row, column=9, value=f"=H{row}/($C$5*1000)*100")
        fmt_number(ws, row, 9, "0.00")

    last_data_row = first_data_row + len(oxides) - 1
    alt_shading(ws, first_data_row, last_data_row, len(headers))
    apply_borders(ws, OX_TABLE_ROW, last_data_row, len(headers))

    # Row/col references for INDEX/MATCH formulas below
    # Oxide name range: A19:A23 (for MATCH against C15 dropdown)
    ox_range = f"$A${first_data_row}:$A${last_data_row}"
    # Density col (E), Recommended col (H)
    rho_range = f"$E${first_data_row}:$E${last_data_row}"
    rec_range = f"$H${first_data_row}:$H${last_data_row}"

    # ── Section 3: Particle Size Sweep (dropdown-driven) ──
    ps_start = last_data_row + 3  # header row
    # Title references the selected oxide via formula
    ws.cell(row=ps_start - 1, column=1,
            value="PARTICLE SIZE SWEEP").font = section_font
    ws.cell(row=ps_start - 1, column=4,
            value='="\u2190 for "&C15&" (change dropdown in C15 to switch)"')
    ws.cell(row=ps_start - 1, column=4).font = note_font

    # Lookup cells for the selected oxide (hidden math, visible for audit)
    # Put them in cols G-I of the section title row so they're visible but
    # out of the way
    ws.cell(row=ps_start - 1, column=7,
            value="Selected \u03c1:").font = note_font
    ws.cell(row=ps_start - 1, column=8,
            value=f"=INDEX({rho_range},MATCH($C$15,{ox_range},0))")
    fmt_number(ws, ps_start - 1, 8, "0")
    ws.cell(row=ps_start - 1, column=9,
            value="kg/m\u00b3").font = note_font

    # Reference cells for formulas below
    rho_cell = f"$H${ps_start - 1}"   # looked-up density
    rec_cell = f"INDEX({rec_range},MATCH($C$15,{ox_range},0))"

    ps_headers = ["Radius (\u03bcm)", "N particles",
                  "Surface area (cm\u00b2)", "t_diff (s)", "t_diff (min)",
                  "Rec. mass (g)"]
    for ci, h in enumerate(ps_headers, 1):
        ws.cell(row=ps_start, column=ci, value=h)
    style_header(ws, ps_start, len(ps_headers))

    radii = [10, 25, 50, 100, 250, 500]
    for i, rad in enumerate(radii):
        row = ps_start + 1 + i
        ws.cell(row=row, column=1, value=rad)
        # F: Rec mass (repeated for clarity — from INDEX/MATCH)
        ws.cell(row=row, column=6,
                value=f"={rec_cell}")
        fmt_number(ws, row, 6, "0.00")
        # B: N = rec_mass_g / (rho * 4/3*pi*r^3 * 1000)
        ws.cell(row=row, column=2,
                value=f"=F{row}/({rho_cell}*4/3*PI()*(A{row}*1E-6)^3*1000)")
        fmt_number(ws, row, 2, "0.00E+00")
        # C: Area = N * 4*pi*r^2 * 1e4
        ws.cell(row=row, column=3,
                value=f"=B{row}*4*PI()*(A{row}*1E-6)^2*1E4")
        fmt_number(ws, row, 3, "0.0")
        # D: t_diff = r^2 / (2*D_Cu)
        ws.cell(row=row, column=4,
                value=f"=(A{row}*1E-6)^2/(2*$C$10)")
        fmt_number(ws, row, 4, "0.0")
        # E: t_diff in minutes
        ws.cell(row=row, column=5,
                value=f"=D{row}/60")
        fmt_number(ws, row, 5, "0.1")

    ps_end = ps_start + len(radii)
    alt_shading(ws, ps_start + 1, ps_end, len(ps_headers))
    apply_borders(ws, ps_start, ps_end, len(ps_headers))

    # ── Key insight note ──
    note_row = ps_end + 2
    ws.cell(row=note_row, column=1,
            value="KEY INSIGHT:").font = Font(bold=True, size=10,
                                              color="2F5496")
    ws.cell(row=note_row + 1, column=1,
            value="D_Cu = 9.63\u00d710\u207b\u00b9\u2070 m\u00b2/s means "
                  "Cu reaches any 100\u03bcm particle in ~5 seconds. "
                  "Diffusion is NOT the bottleneck; the chemical "
                  "reaction at the particle surface is rate-limiting."
            ).font = note_font
    ws.merge_cells(start_row=note_row + 1, start_column=1,
                   end_row=note_row + 1, end_column=9)

    # ── Section 4: "What-If" Reverse Calculator ──
    wi_start = note_row + 4
    ws.cell(row=wi_start, column=1,
            value="WHAT-IF: I HAVE THIS MUCH OXIDE").font = section_font
    ws.cell(row=wi_start + 1, column=1,
            value="Enter the oxide mass you actually have. "
                  "Formulas compute how much Cu it can remove."
            ).font = note_font
    ws.merge_cells(start_row=wi_start + 1, start_column=1,
                   end_row=wi_start + 1, end_column=6)

    # Input cell: oxide mass available
    wi_input_row = wi_start + 3
    ws.cell(row=wi_input_row, column=1,
            value="Oxide mass available").font = Font(size=10)
    ws.cell(row=wi_input_row, column=2, value="m_oxide")
    c_avail = ws.cell(row=wi_input_row, column=3, value=2.0)
    c_avail.fill = DROPDOWN_FILL
    c_avail.font = DROPDOWN_FONT
    c_avail.border = DROPDOWN_BORDER
    ws.cell(row=wi_input_row, column=4,
            value="g (EDIT THIS \u2192 results update below)")

    # Derived values: how much Cu this oxide can capture
    # Uses the selected oxide from dropdown C15
    # cu_per_mol from lookup: INDEX(F19:F23, MATCH(C15, A19:A23, 0))
    cu_per_mol_range = f"$F${first_data_row}:$F${last_data_row}"
    mw_range = f"$D${first_data_row}:$D${last_data_row}"

    wi_r = wi_input_row + 1
    ws.cell(row=wi_r, column=1, value="MW of selected oxide")
    ws.cell(row=wi_r, column=3,
            value=f"=INDEX({mw_range},MATCH($C$15,{ox_range},0))")
    fmt_number(ws, wi_r, 3, "0.00")
    ws.cell(row=wi_r, column=4, value="g/mol")

    wi_r += 1
    ws.cell(row=wi_r, column=1, value="Cu captured per mol oxide")
    ws.cell(row=wi_r, column=3,
            value=f"=INDEX({cu_per_mol_range},MATCH($C$15,{ox_range},0))")
    fmt_number(ws, wi_r, 3, "0.0")
    ws.cell(row=wi_r, column=4, value="mol Cu / mol oxide")

    wi_r += 1
    ws.cell(row=wi_r, column=1, value="Oxide moles available")
    ws.cell(row=wi_r, column=3,
            value=f"=C{wi_input_row}/C{wi_input_row+1}")
    fmt_number(ws, wi_r, 3, "0.0000")
    ws.cell(row=wi_r, column=4, value="mol")

    wi_r += 1
    ws.cell(row=wi_r, column=1, value="Cu moles removable (stoich)")
    ws.cell(row=wi_r, column=3,
            value=f"=C{wi_r-1}*C{wi_r-2}")
    fmt_number(ws, wi_r, 3, "0.0000")
    ws.cell(row=wi_r, column=4, value="mol Cu")

    wi_r += 1
    ws.cell(row=wi_r, column=1, value="Cu mass removable")
    ws.cell(row=wi_r, column=3,
            value=f"=C{wi_r-1}*63.546")
    fmt_number(ws, wi_r, 3, "0.00")
    ws.cell(row=wi_r, column=4, value="g Cu")

    wi_r += 1
    ws.cell(row=wi_r, column=1, value="Cu reduction in wt%")
    ws.cell(row=wi_r, column=3,
            value=f"=C{wi_r-1}/($C$5*1000)*100")
    fmt_number(ws, wi_r, 3, "0.000")
    ws.cell(row=wi_r, column=4,
            value="wt% Cu removed from steel")

    wi_r += 1
    ws.cell(row=wi_r, column=1,
            value="Final Cu (if stoich. complete)").font = \
        Font(bold=True, size=10)
    ws.cell(row=wi_r, column=3,
            value=f"=MAX(0,$C$6-C{wi_r-1})")
    ws.cell(row=wi_r, column=3).font = Font(bold=True, size=10)
    fmt_number(ws, wi_r, 3, "0.000")
    ws.cell(row=wi_r, column=4,
            value="wt% Cu remaining (assuming 100% conversion)")

    # Excess ratio indicator
    wi_r += 1
    ws.cell(row=wi_r, column=1, value="Excess ratio vs need")
    ws.cell(row=wi_r, column=3,
            value=f"=IF($C$12>0,C{wi_input_row}/"
                  f"INDEX($G${first_data_row}:$G${last_data_row},"
                  f"MATCH($C$15,{ox_range},0)),\"N/A\")")
    fmt_number(ws, wi_r, 3, "0.0")
    ws.cell(row=wi_r, column=4,
            value='x stoichiometric (need >= 1.0 for full removal)')

    alt_shading(ws, wi_input_row, wi_r, 4)
    apply_borders(ws, wi_input_row, wi_r, 4)

    # Formatting
    freeze_and_color(ws, BLUE_TAB, freeze_row=4)
    auto_width(ws, 9)

    print(f"  Mass_Balance: parameters + {len(oxides)} oxides + "
          f"{len(radii)}-point particle sweep (dropdown-driven)")
    return ws


def build_normalization(wb):
    ws = wb.create_sheet("Normalization")

    fields_ox, rows_ox = load_csv(CSV_FILES["ox_gibbs"])
    if not fields_ox:
        ws.cell(row=1, column=1, value="OX_Gibbs data not available")
        freeze_and_color(ws, BLUE_TAB)
        return ws

    def col_letter_for(header):
        try:
            return get_column_letter(fields_ox.index(header) + 1)
        except ValueError:
            return None

    # Find T=1800K row in OX_Gibbs (accounting for intro offset)
    ox_row_1800 = None
    for i, r in enumerate(rows_ox):
        if abs(float(r["T_K"]) - 1800) < 1:
            ox_row_1800 = i + 2 + INTRO_ROWS
            break
    if ox_row_1800 is None:
        ox_row_1800 = len(rows_ox) + 1 + INTRO_ROWS

    # G_O2 from CuFe2O4_Raw (col H, row for T=1800, with intro offset)
    _, cufe_rows = load_csv(CSV_FILES["cufe2o4"])
    cufe_row_1800 = None
    for i, r in enumerate(cufe_rows):
        if abs(float(r["T_K"]) - 1800) < 1:
            cufe_row_1800 = i + 2 + INTRO_ROWS
            break
    g_o2_ref = (f"=CuFe2O4_Raw!H{cufe_row_1800}"
                if cufe_row_1800 else "N/A")

    # ── Intro card ──
    _write_card_line(ws, 1, 1, 9,
                     "NORMALIZATION \u2014 How TC-Python GM values become "
                     "\u0394G\u1da0 (kJ/mol O\u2082)",
                     font=Font(bold=True, size=11, color="2F5496"))
    _write_card_line(ws, 2, 1, 9,
                     "Section 1: Binary oxide stoichiometry (17 oxides). "
                     "Section 2: Ternary product stoichiometry (18 products). "
                     "Sections 3-4: Worked examples with traceable formulas.",
                     font=Font(size=9, italic=True, color="666666"))

    r_pos = ox_row_1800  # row in OX_Gibbs (with intro offset)

    # ==================================================================
    # SECTION 1: Binary Oxide Stoichiometry Table
    # ==================================================================
    sec1 = 4
    _write_card_line(ws, sec1, 1, 9,
                     "SECTION 1: BINARY OXIDE NORMALIZATION",
                     font=Font(bold=True, size=11))
    ws.cell(row=sec1 + 1, column=1,
            value="TC-Python reports GM = Gibbs energy per mole of ATOMS. "
                  "Pipeline: GM \u00d7 atoms/formula \u2192 G(oxide); "
                  "then \u0394G\u1da0 = G(oxide) \u2212 G(metals) \u2212 G(O\u2082); "
                  "finally \u0394G\u1da0 / n_O\u2082 \u2192 kJ/mol O\u2082 "
                  "for Ellingham comparison."
    ).font = Font(size=9, color="333333")
    ws.merge_cells(start_row=sec1 + 1, start_column=1,
                   end_row=sec1 + 1, end_column=9)

    bin_hdr = sec1 + 3
    bin_headers = ["Oxide", "Atoms/Formula", "Balanced Reaction",
                   "oxide/O\u2082", "GM @ 1800 K (J/mol-at)",
                   "\u0394G per mol O\u2082 (kJ)"]
    bin_ncols = len(bin_headers)
    for c, h in enumerate(bin_headers, 1):
        ws.cell(row=bin_hdr, column=c, value=h)
    style_header(ws, bin_hdr, bin_ncols)

    for r_idx, (oxide, atoms, reaction, ratio) in enumerate(NORM_TABLE):
        r = bin_hdr + 1 + r_idx
        ws.cell(row=r, column=1, value=oxide)
        ws.cell(row=r, column=2, value=atoms)
        ws.cell(row=r, column=3, value=reaction)
        ws.cell(row=r, column=4, value=ratio)
        fmt_number(ws, r, 4, "0.000")
        # GM from OX_Gibbs
        gm_col = col_letter_for(f"GM_{oxide}")
        if gm_col:
            ws.cell(row=r, column=5,
                    value=f"=OX_Gibbs!{gm_col}{r_pos}")
            fmt_number(ws, r, 5, "#,##0")
        else:
            ws.cell(row=r, column=5, value="N/A")
        # dG per mol O2 from OX_Gibbs (pre-computed)
        dg_col = col_letter_for(f"dG_{oxide}_per_O2")
        if dg_col:
            ws.cell(row=r, column=6,
                    value=f"=OX_Gibbs!{dg_col}{r_pos}/1000")
            fmt_number(ws, r, 6, "#,##0.0")
        else:
            ws.cell(row=r, column=6, value="N/A")

    bin_last = bin_hdr + len(NORM_TABLE)
    alt_shading(ws, bin_hdr + 1, bin_last, bin_ncols)
    center_all(ws, bin_hdr, bin_last, bin_ncols)
    # Left-align reaction column
    for rr in range(bin_hdr + 1, bin_last + 1):
        ws.cell(row=rr, column=3).alignment = Alignment(
            horizontal="left", wrap_text=True)
    apply_borders(ws, bin_hdr, bin_last, bin_ncols)

    # ── Bar chart: ΔG per mol O₂ for all binary oxides at 1800 K ──
    # Placed beside the table (col H) aligned with the header row
    from openpyxl.drawing.fill import PatternFillProperties, ColorChoice
    from openpyxl.chart.label import DataLabelList

    bar = BarChart()
    bar.type = "bar"           # horizontal bars
    bar.grouping = "clustered"
    bar.title = "ΔG° per mol O₂ at 1800 K"
    # For type="bar" (horizontal): x_axis=CATEGORY (vertical), y_axis=VALUE (horizontal)
    bar.x_axis.title = None
    bar.y_axis.title = "kJ/mol O₂"
    bar.width = 28
    bar.height = 28
    bar.style = 10

    bar_data = Reference(ws, min_col=6, min_row=bin_hdr,
                         max_row=bin_last)
    bar_cats = Reference(ws, min_col=1, min_row=bin_hdr + 1,
                         max_row=bin_last)
    bar.add_data(bar_data, titles_from_data=True)
    bar.set_categories(bar_cats)
    bar.legend = None

    # Blue fill — use pct90 pattern (near-solid) as workaround
    s_bar = bar.series[0]
    s_bar.graphicalProperties.pattFill = PatternFillProperties(
        prst="pct90",
        fgClr=ColorChoice(srgbClr="0077BB"),
        bgClr=ColorChoice(srgbClr="0077BB"),
    )
    s_bar.graphicalProperties.line.solidFill = "005588"
    s_bar.graphicalProperties.line.width = 6350

    # Data labels — numeric values at the tip (left end) of each bar
    s_bar.dLbls = DataLabelList()
    s_bar.dLbls.showVal = True
    s_bar.dLbls.showCatName = False
    s_bar.dLbls.showSerName = False
    s_bar.dLbls.showLegendKey = False
    s_bar.dLbls.numFmt = "#,##0"
    s_bar.dLbls.dLblPos = "outEnd"   # at the tip of each bar

    # Light gray gridlines
    grid_sp_bar = GraphicalProperties(
        ln=LineProperties(solidFill="D0D0D0", w=6350))
    bar.x_axis.majorGridlines = ChartLines(spPr=grid_sp_bar)
    bar.y_axis.majorGridlines = ChartLines(spPr=grid_sp_bar)
    # x_axis = CATEGORY axis (vertical, oxide names)
    bar.x_axis.delete = False
    bar.x_axis.tickLblPos = "high"   # oxide names on RIGHT
    bar.x_axis.crosses = "max"       # category crosses at x=0 (right side)
    # y_axis = VALUE axis (horizontal, numbers)
    bar.y_axis.delete = False
    bar.y_axis.tickLblPos = "low"    # numbers at BOTTOM

    ws.add_chart(bar, f"H{bin_hdr}")

    # ==================================================================
    # SECTION 2: Ternary Product Stoichiometry Table
    # ==================================================================
    sec2 = bin_last + 2
    _write_card_line(ws, sec2, 1, 9,
                     "SECTION 2: TERNARY PRODUCT STOICHIOMETRY",
                     font=Font(bold=True, size=11))
    ws.cell(row=sec2 + 1, column=1,
            value="Ternary reaction: Cu + MO\u2093 + O\u2082 \u2192 CuMO\u1d67. "
                  "If \u0394G < 0 the oxide captures copper into a ternary "
                  "compound (or slag dissolution at steelmaking T)."
    ).font = Font(size=9, color="333333")
    ws.merge_cells(start_row=sec2 + 1, start_column=1,
                   end_row=sec2 + 1, end_column=9)

    tern_hdr = sec2 + 3
    tern_headers = ["Product", "Name", "Reaction", "Parent Oxide",
                    "n_Cu", "n_oxide", "n_O\u2082",
                    "Product Atoms", "Category"]
    tern_ncols = len(tern_headers)
    for c, h in enumerate(tern_headers, 1):
        ws.cell(row=tern_hdr, column=c, value=h)
    style_header(ws, tern_hdr, tern_ncols)

    for r_idx, rxn_data in enumerate(TERNARY_REACTIONS):
        prod, name, rxn, oxide, n_cu, n_ox, n_o2, p_at, cat = rxn_data
        r = tern_hdr + 1 + r_idx
        for c, val in enumerate(
                [prod, name, rxn, oxide, n_cu, n_ox, n_o2, p_at, cat], 1):
            ws.cell(row=r, column=c, value=val)

    tern_last = tern_hdr + len(TERNARY_REACTIONS)
    alt_shading(ws, tern_hdr + 1, tern_last, tern_ncols)
    center_all(ws, tern_hdr, tern_last, tern_ncols)
    for rr in range(tern_hdr + 1, tern_last + 1):
        ws.cell(row=rr, column=3).alignment = Alignment(
            horizontal="left", wrap_text=True)
    apply_borders(ws, tern_hdr, tern_last, tern_ncols)

    # ==================================================================
    # SECTION 3: Worked Example — Al2O3
    # ==================================================================
    ex1_start = tern_last + 2
    _write_card_line(ws, ex1_start, 1, 5,
                     "SECTION 3: WORKED EXAMPLE \u2014 Al\u2082O\u2083 (corundum)",
                     font=Font(bold=True, size=11))
    ws.cell(row=ex1_start + 1, column=1,
            value="2Al + 1.5 O\u2082 \u2192 Al\u2082O\u2083. Every cell "
                  "traces to OX_Gibbs or CuFe2O4_Raw."
    ).font = Font(size=9, color="333333")
    ws.merge_cells(start_row=ex1_start + 1, start_column=1,
                   end_row=ex1_start + 1, end_column=5)

    ex1_hdr = ex1_start + 3
    norm_headers = ["Step", "Description", "Value", "Unit"]
    norm_ncols = len(norm_headers)
    for c, h in enumerate(norm_headers, 1):
        ws.cell(row=ex1_hdr, column=c, value=h)
    style_header(ws, ex1_hdr, norm_ncols)

    gm_al = col_letter_for("GM_Al2O3")
    gm_al_metal = col_letter_for("G_metal_Al2O3")
    dg_al_check = col_letter_for("dG_Al2O3_per_O2")
    R = ex1_hdr

    steps_al = [
        ("1", "GM per atom from TC-Python",
         f"=OX_Gibbs!{gm_al}{r_pos}" if gm_al else "N/A",
         "J/mol-atoms"),
        ("2", "Atoms per formula unit (Al\u2082O\u2083 = 5)",
         5, "atoms"),
        ("3", "G per formula unit = GM \u00d7 atoms/formula",
         f"=C{R+1}*C{R+2}" if gm_al else "N/A", "J/mol"),
        ("4", "G of pure Al metal reference",
         f"=OX_Gibbs!{gm_al_metal}{r_pos}" if gm_al_metal else "N/A",
         "J/mol"),
        ("5", "G of O\u2082 gas reference (from CuFe2O4_Raw)",
         g_o2_ref, "J/mol"),
        ("6", "\u0394G\u1da0 = G(oxide) \u2212 2\u00b7G(metal) \u2212 "
              "1.5\u00b7G(O\u2082)",
         f"=C{R+3}-2*C{R+4}-1.5*C{R+5}", "J/mol"),
        ("7", "\u0394G\u1da0 per mol O\u2082 = \u0394G\u1da0 / 1.5 / 1000",
         f"=C{R+6}/1.5/1000", "kJ/mol O\u2082"),
        ("8", "TC-Python pre-computed (verification)",
         f"=OX_Gibbs!{dg_al_check}{r_pos}/1000" if dg_al_check else "N/A",
         "kJ/mol O\u2082"),
    ]

    for j, (step, desc, formula, unit) in enumerate(steps_al):
        row = R + 1 + j
        ws.cell(row=row, column=1, value=step)
        ws.cell(row=row, column=2, value=desc)
        ws.cell(row=row, column=3, value=formula)
        ws.cell(row=row, column=4, value=unit)
        if "J/mol" in unit:
            fmt_number(ws, row, 3, "#,##0.0")

    ex1_last = R + len(steps_al)
    alt_shading(ws, R + 1, ex1_last, norm_ncols)
    apply_borders(ws, R, ex1_last, norm_ncols)

    # ==================================================================
    # SECTION 4: Worked Example — Cu2O
    # ==================================================================
    ex2_start = ex1_last + 2
    _write_card_line(ws, ex2_start, 1, 5,
                     "SECTION 4: WORKED EXAMPLE \u2014 Cu\u2082O (cuprite)",
                     font=Font(bold=True, size=11))
    ws.cell(row=ex2_start + 1, column=1,
            value="4Cu + O\u2082 \u2192 2Cu\u2082O. This is the reference "
                  "oxide \u2014 all screening compares against Cu\u2082O."
    ).font = Font(size=9, color="333333")
    ws.merge_cells(start_row=ex2_start + 1, start_column=1,
                   end_row=ex2_start + 1, end_column=5)

    ex2_hdr = ex2_start + 3
    for c, h in enumerate(norm_headers, 1):
        ws.cell(row=ex2_hdr, column=c, value=h)
    style_header(ws, ex2_hdr, norm_ncols)

    gm_cu = col_letter_for("GM_Cu2O")
    gm_cu_metal = col_letter_for("G_metal_Cu2O")
    dg_cu_check = col_letter_for("dG_Cu2O_per_O2")
    R2 = ex2_hdr

    steps_cu = [
        ("1", "GM per atom from TC-Python",
         f"=OX_Gibbs!{gm_cu}{r_pos}" if gm_cu else "N/A",
         "J/mol-atoms"),
        ("2", "Atoms per formula unit (Cu\u2082O = 3)",
         3, "atoms"),
        ("3", "G per formula unit = GM \u00d7 atoms/formula",
         f"=C{R2+1}*C{R2+2}" if gm_cu else "N/A", "J/mol"),
        ("4", "G of pure Cu metal reference",
         f"=OX_Gibbs!{gm_cu_metal}{r_pos}" if gm_cu_metal else "N/A",
         "J/mol"),
        ("5", "G of O\u2082 gas reference (from CuFe2O4_Raw)",
         g_o2_ref, "J/mol"),
        ("6", "\u0394G\u1da0 = G(oxide) \u2212 2\u00b7G(metal) \u2212 "
              "0.5\u00b7G(O\u2082)",
         f"=C{R2+3}-2*C{R2+4}-0.5*C{R2+5}", "J/mol"),
        ("7", "\u0394G\u1da0 per mol O\u2082 = \u0394G\u1da0 / 0.5 / 1000",
         f"=C{R2+6}/0.5/1000", "kJ/mol O\u2082"),
        ("8", "TC-Python pre-computed (verification)",
         f"=OX_Gibbs!{dg_cu_check}{r_pos}/1000" if dg_cu_check else "N/A",
         "kJ/mol O\u2082"),
    ]

    for j, (step, desc, formula, unit) in enumerate(steps_cu):
        row = R2 + 1 + j
        ws.cell(row=row, column=1, value=step)
        ws.cell(row=row, column=2, value=desc)
        ws.cell(row=row, column=3, value=formula)
        ws.cell(row=row, column=4, value=unit)
        if "J/mol" in unit:
            fmt_number(ws, row, 3, "#,##0.0")

    ex2_last = R2 + len(steps_cu)
    alt_shading(ws, R2 + 1, ex2_last, norm_ncols)
    apply_borders(ws, R2, ex2_last, norm_ncols)

    # Column widths
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 38
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 22
    ws.column_dimensions["F"].width = 22
    ws.column_dimensions["G"].width = 10
    ws.column_dimensions["H"].width = 14
    ws.column_dimensions["I"].width = 14
    freeze_and_color(ws, BLUE_TAB)

    print(f"  Normalization: {len(NORM_TABLE)} binary + "
          f"{len(TERNARY_REACTIONS)} ternary + 2 worked examples "
          f"(OX_Gibbs row {r_pos})")
    return ws


# =====================================================================
# main
# =====================================================================
def main():
    print("=" * 70)
    print("Building Cu_Removal_Unified.xlsx")
    print("  6 analysis (blue) + 2 results (green) + 8 raw/backing (gray) = 16 tabs")
    print("=" * 70)

    wb = openpyxl.Workbook()

    # Blue analysis tabs (ordered by audience priority)
    print("\n--- Analysis Tabs (blue) ---")
    build_dashboard(wb)
    build_screening_summary(wb)
    build_mass_balance(wb)
    build_activity_and_slag(wb)
    build_cufe2o4_deepdive(wb)
    build_normalization(wb)

    # Green results tabs (Screening_Table + Activity_Corrected)
    # Ternary_Verdicts demoted to gray (backing data for Screening_Summary)
    print("\n--- Results Tabs (green) + Raw/Backing (gray) ---")
    build_results_tabs(wb)

    # Gray raw data tabs
    build_raw_tabs(wb)

    wb.save(OUTPUT_FILE)
    print(f"\nSaved: {OUTPUT_FILE}")
    print(f"  {len(wb.sheetnames)} tabs: {', '.join(wb.sheetnames)}")
    print("=" * 70)


if __name__ == "__main__":
    main()
