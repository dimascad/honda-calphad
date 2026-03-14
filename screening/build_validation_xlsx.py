"""
Build Validation_Results.xlsx with TRACEABLE CELL FORMULAS.

Every derived value is an Excel formula (e.g., =C2-D2) so anyone can
click a cell and see the arithmetic.  Raw TC-Python values are written
as numbers; anything computed FROM those values is a formula string.

Sheets:
  1. dG_vs_T         — pivoted dG data for top 6 products
  2. CuFe2O4_Decomposition — original vs alternative dG with formula columns
  3. Activity_Correction — dilute Cu penalty with all formulas visible
  4. Candidate_Ranking — summary with cross-sheet references
  5. Slag_Effects     — slag composition data + % change formulas
  6. Cu_Activity      — raw activity data
  7. Data_Sources     — provenance documentation
"""

import csv
from pathlib import Path

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers,
)
from openpyxl.utils import get_column_letter

# ── paths ──────────────────────────────────────────────────────────────
SCRIPT_DIR = Path(__file__).parent
DATA_DIR = SCRIPT_DIR.parent / "data" / "tcpython" / "raw"
OUTPUT_FILE = SCRIPT_DIR / "Validation_Results.xlsx"

PROC_DIR = SCRIPT_DIR.parent / "data" / "tcpython" / "processed"

CSV_FILES = {
    "dG_top6":      DATA_DIR / "dG_vs_T_top6.csv",
    "cufe2o4":      DATA_DIR / "cufe2o4_alternative_reaction.csv",
    "decomp":       SCRIPT_DIR / "cufe2o4_decomposition_results.csv",
    "slag":         DATA_DIR / "slag_composition_effects.csv",
    "activity":     DATA_DIR / "cu_activity_vs_oxide.csv",
    "corrected":    PROC_DIR / "activity_corrected_dG.csv",
}

# ── IEEE formatting constants (per CLAUDE.md) ─────────────────────────
HEADER_FILL = PatternFill(start_color="595959", end_color="595959",
                          fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
WHITE_FILL  = PatternFill(start_color="FFFFFF", end_color="FFFFFF",
                          fill_type="solid")
GRAY_FILL   = PatternFill(start_color="E0E0E0", end_color="E0E0E0",
                          fill_type="solid")

THICK = Side(style="medium")
THIN  = Side(style="thin")


# ── helpers ────────────────────────────────────────────────────────────
def load_csv(path):
    """Return (fieldnames, list[dict])."""
    if not path.exists():
        return None, []
    with open(path) as f:
        reader = csv.DictReader(f)
        fieldnames = reader.fieldnames
        rows = list(reader)
    return fieldnames, rows


def _try_float(v):
    """Convert string to float if possible, else return string."""
    try:
        return float(v)
    except (TypeError, ValueError):
        return v


def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)


def apply_row_fill(ws, row, ncols, fill):
    for c in range(1, ncols + 1):
        ws.cell(row=row, column=c).fill = fill


def alt_shading(ws, start_row, end_row, ncols, group_size=1):
    """Alternate white/gray fills by group."""
    for r in range(start_row, end_row + 1):
        grp = (r - start_row) // group_size
        fill = WHITE_FILL if grp % 2 == 0 else GRAY_FILL
        apply_row_fill(ws, r, ncols, fill)


def apply_borders(ws, start_row, end_row, ncols):
    """Thick outer frame + bottom-of-header; thin internal."""
    for r in range(start_row, end_row + 1):
        for c in range(1, ncols + 1):
            top = THICK if r == start_row else THIN
            bot = THICK if r == end_row else THIN
            left = THICK if c == 1 else THIN
            right = THICK if c == ncols else THIN
            # Header bottom gets thick
            if r == start_row:
                bot = THICK
            ws.cell(row=r, column=c).border = Border(
                top=top, bottom=bot, left=left, right=right
            )


def auto_width(ws, ncols, min_w=8, max_w=30):
    for c in range(1, ncols + 1):
        best = min_w
        for cell in ws.iter_rows(min_col=c, max_col=c, values_only=False):
            for ce in cell:
                if ce.value is not None:
                    best = max(best, len(str(ce.value)) + 2)
        ws.column_dimensions[get_column_letter(c)].width = min(best, max_w)


def center_all(ws, start_row, end_row, ncols):
    for r in range(start_row, end_row + 1):
        for c in range(1, ncols + 1):
            ws.cell(row=r, column=c).alignment = Alignment(horizontal="center")


def fmt_number(ws, row, col, fmt_str):
    ws.cell(row=row, column=col).number_format = fmt_str


# =====================================================================
# Sheet 1: dG_vs_T  (pivoted — one column per product)
# =====================================================================
def build_dG_vs_T(wb):
    _, rows = load_csv(CSV_FILES["dG_top6"])
    if not rows:
        ws = wb.active
        ws.title = "dG_vs_T"
        ws.cell(row=1, column=1, value="Data not available")
        return ws

    # Pivot: collect unique temps and products in order
    temps = []
    seen_t = set()
    products = []
    seen_p = set()
    data = {}  # (T_K, product) -> dG_kJ
    for r in rows:
        tk = int(r["T_K"])
        prod = r["product"]
        if tk not in seen_t:
            temps.append(tk)
            seen_t.add(tk)
        if prod not in seen_p:
            products.append(prod)
            seen_p.add(prod)
        dg_str = r["dG_rxn_system_kJ"].strip()
        if dg_str:
            data[(tk, prod)] = float(dg_str)

    ws = wb.active
    ws.title = "dG_vs_T"

    ncols = 2 + len(products)  # T(K), T(C), then one per product

    # Header row
    ws.cell(row=1, column=1, value="T (K)")
    ws.cell(row=1, column=2, value="T (\u00b0C)")
    for i, prod in enumerate(products):
        ws.cell(row=1, column=3 + i, value=f"dG {prod} (kJ)")
    style_header(ws, 1, ncols)

    # Data rows
    for row_idx, tk in enumerate(sorted(temps), start=2):
        # Column A: T_K (raw number)
        ws.cell(row=row_idx, column=1, value=tk)
        # Column B: T_C as FORMULA  =A{row}-273.15
        ws.cell(row=row_idx, column=2, value=f"=A{row_idx}-273.15")
        fmt_number(ws, row_idx, 2, "0.00")
        # Columns C onward: raw dG values
        for i, prod in enumerate(products):
            val = data.get((tk, prod))
            col = 3 + i
            if val is not None:
                ws.cell(row=row_idx, column=col, value=val)
                fmt_number(ws, row_idx, col, "+0.00;-0.00;0.00")

    last_data_row = 1 + len(sorted(set(temps)))

    # MIN row — formula per product column
    min_row = last_data_row + 1
    ws.cell(row=min_row, column=1, value="MIN")
    ws.cell(row=min_row, column=1).font = Font(bold=True, size=10)
    ws.cell(row=min_row, column=2, value="")
    for i in range(len(products)):
        col_letter = get_column_letter(3 + i)
        ws.cell(
            row=min_row, column=3 + i,
            value=f"=MIN({col_letter}2:{col_letter}{last_data_row})"
        )
        fmt_number(ws, min_row, 3 + i, "+0.00;-0.00;0.00")
    apply_row_fill(ws, min_row, ncols, GRAY_FILL)
    ws.cell(row=min_row, column=1).fill = GRAY_FILL

    # Formatting
    alt_shading(ws, 2, last_data_row, ncols)
    center_all(ws, 1, min_row, ncols)
    apply_borders(ws, 1, min_row, ncols)
    auto_width(ws, ncols)

    print(f"  Sheet 1: dG_vs_T ({len(temps)} temps x {len(products)} products)")
    return ws


# =====================================================================
# Sheet 2: CuFe2O4_Decomposition
# =====================================================================
def build_cufe2o4_decomp(wb):
    ws = wb.create_sheet("CuFe2O4_Decomposition")

    # Prefer the alternative_reaction CSV (has the raw GM/G values)
    _, rows = load_csv(CSV_FILES["cufe2o4"])
    if not rows:
        # Fall back to pre-computed decomposition CSV
        _, rows = load_csv(CSV_FILES["decomp"])
        if not rows:
            ws.cell(row=1, column=1, value="Data not available")
            return ws
        # Write pre-computed with formulas where possible
        _build_decomp_from_precomputed(ws, rows)
        return ws

    # ── Column layout ──
    # A: T(K)              raw
    # B: T(°C)             formula =A-273.15
    # C: dG_original (kJ)  raw   (full reaction Cu+2FeO+O2->CuFe2O4)
    # D: dG_alternative (kJ) raw (Cu-capture-only reaction)
    # E: dG_Fe_oxidation   FORMULA =C-D
    # F: Cu_capture_%      FORMULA =D/C*100
    headers = [
        "T (K)",
        "T (\u00b0C)",
        "dG_original (kJ)",
        "dG_alternative (kJ)",
        "dG_Fe_oxidation (kJ)",
        "Cu_capture (%)",
    ]
    ncols = len(headers)
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header(ws, 1, ncols)

    for i, r in enumerate(rows, start=2):
        # A: T_K raw
        ws.cell(row=i, column=1, value=_try_float(r["T_K"]))
        # B: T_C formula
        ws.cell(row=i, column=2, value=f"=A{i}-273.15")
        fmt_number(ws, i, 2, "0.00")
        # C: dG_original raw
        ws.cell(row=i, column=3, value=_try_float(r["dG_original_kJ"]))
        fmt_number(ws, i, 3, "+0.00;-0.00;0.00")
        # D: dG_alternative raw
        ws.cell(row=i, column=4, value=_try_float(r["dG_alternative_kJ"]))
        fmt_number(ws, i, 4, "+0.00;-0.00;0.00")
        # E: Fe oxidation = FORMULA  =C{i}-D{i}
        ws.cell(row=i, column=5, value=f"=C{i}-D{i}")
        fmt_number(ws, i, 5, "+0.00;-0.00;0.00")
        # F: Cu capture % = FORMULA  =D{i}/C{i}*100
        ws.cell(row=i, column=6, value=f"=D{i}/C{i}*100")
        fmt_number(ws, i, 6, "0.0")

    last_row = 1 + len(rows)
    alt_shading(ws, 2, last_row, ncols)
    center_all(ws, 1, last_row, ncols)
    apply_borders(ws, 1, last_row, ncols)
    auto_width(ws, ncols)

    print(f"  Sheet 2: CuFe2O4_Decomposition ({len(rows)} rows, formula cols E & F)")
    return ws


def _build_decomp_from_precomputed(ws, rows):
    """Fallback: use pre-computed CSV but still add formula columns."""
    headers = [
        "T (K)", "T (\u00b0C)",
        "dG_alternative (kJ)", "dG_original (kJ)",
        "dG_Fe_oxidation (kJ)", "Cu_capture (%)",
    ]
    ncols = len(headers)
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header(ws, 1, ncols)

    for i, r in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=_try_float(r["T_K"]))
        ws.cell(row=i, column=2, value=f"=A{i}-273.15")
        fmt_number(ws, i, 2, "0.00")
        ws.cell(row=i, column=3, value=_try_float(r["dG_alternative_kJ"]))
        fmt_number(ws, i, 3, "+0.00;-0.00;0.00")
        ws.cell(row=i, column=4, value=_try_float(r["dG_original_kJ"]))
        fmt_number(ws, i, 4, "+0.00;-0.00;0.00")
        # E: Fe oxidation = original - alternative = D - C
        ws.cell(row=i, column=5, value=f"=D{i}-C{i}")
        fmt_number(ws, i, 5, "+0.00;-0.00;0.00")
        # F: Cu capture % = alternative / original * 100 = C / D * 100
        ws.cell(row=i, column=6, value=f"=C{i}/D{i}*100")
        fmt_number(ws, i, 6, "0.0")

    last_row = 1 + len(rows)
    alt_shading(ws, 2, last_row, ncols)
    center_all(ws, 1, last_row, ncols)
    apply_borders(ws, 1, last_row, ncols)
    auto_width(ws, ncols)
    print(f"  Sheet 2: CuFe2O4_Decomposition ({len(rows)} rows, fallback CSV)")


# =====================================================================
# Sheet 3: Activity_Correction  (ALL formulas — fully traceable)
# =====================================================================
def build_activity_correction(wb):
    ws = wb.create_sheet("Activity_Correction")

    # ── Section 1: Input parameters (named cells at top) ──
    param_font = Font(bold=True, size=10)
    note_font = Font(italic=True, size=9, color="666666")

    ws.cell(row=1, column=1, value="ACTIVITY CORRECTION PARAMETERS").font = \
        Font(bold=True, size=12)
    ws.merge_cells("A1:G1")

    ws.cell(row=2, column=1, value="Parameter").font = param_font
    ws.cell(row=2, column=2, value="Symbol").font = param_font
    ws.cell(row=2, column=3, value="Value").font = param_font
    ws.cell(row=2, column=4, value="Unit / Note").font = param_font
    style_header(ws, 2, 4)

    # Row 3: X_Cu
    ws.cell(row=3, column=1, value="Cu mole fraction in steel")
    ws.cell(row=3, column=2, value="X_Cu")
    ws.cell(row=3, column=3, value=0.003)
    fmt_number(ws, 3, 3, "0.0000")
    ws.cell(row=3, column=4, value="~0.3 wt% Cu")

    # Row 4: gamma_Cu
    ws.cell(row=4, column=1, value="Raoultian activity coefficient")
    ws.cell(row=4, column=2, value="\u03b3_Cu")
    ws.cell(row=4, column=3, value=8.5)
    fmt_number(ws, 4, 3, "0.0")
    ws.cell(row=4, column=4, value="Literature range: 5-13")

    # Row 5: a_Cu = gamma * X  (FORMULA)
    ws.cell(row=5, column=1, value="Cu activity in liquid Fe")
    ws.cell(row=5, column=2, value="a_Cu")
    ws.cell(row=5, column=3, value="=C3*C4")
    fmt_number(ws, 5, 3, "0.0000")
    ws.cell(row=5, column=4, value="X_Cu \u00d7 \u03b3_Cu")

    # Row 6: T
    ws.cell(row=6, column=1, value="Temperature")
    ws.cell(row=6, column=2, value="T")
    ws.cell(row=6, column=3, value=1800)
    ws.cell(row=6, column=4, value="K (steelmaking)")

    # Row 7: R
    ws.cell(row=7, column=1, value="Gas constant")
    ws.cell(row=7, column=2, value="R")
    ws.cell(row=7, column=3, value=8.314)
    fmt_number(ws, 7, 3, "0.000")
    ws.cell(row=7, column=4, value="J/(mol\u00b7K)")

    # Row 8: RT*ln(1/a_Cu) = penalty per Cu atom (FORMULA)
    ws.cell(row=8, column=1, value="Penalty per Cu atom consumed")
    ws.cell(row=8, column=2, value="-RT\u00b7ln(a_Cu)")
    ws.cell(row=8, column=3, value="=-C7*C6*LN(C5)/1000")
    fmt_number(ws, 8, 3, "+0.0;-0.0")
    ws.cell(row=8, column=4, value="kJ/mol (positive = less favorable)")

    alt_shading(ws, 3, 8, 4)
    apply_borders(ws, 2, 8, 4)

    # ── Section 2: Product-by-product correction table ──
    TABLE_START = 10
    ws.cell(row=TABLE_START, column=1,
            value="ACTIVITY-CORRECTED \u0394G AT 1800 K").font = \
        Font(bold=True, size=12)
    ws.merge_cells(f"A{TABLE_START}:H{TABLE_START}")

    hdr = TABLE_START + 1
    col_headers = [
        "Product",           # A
        "Parent Oxide",      # B
        "n_Cu",              # C  (atoms consumed)
        "\u0394G\u00b0 (kJ)",  # D  (pure reference, raw from TC-Python)
        "Cu Penalty (kJ)",   # E  = n_Cu * penalty_per_atom  FORMULA
        "\u0394G_eff (kJ)",  # F  = D + E  FORMULA
        "Verdict",           # G  =IF(F<0,"FAVORABLE","UNFAVORABLE")
        "\u0394G per Cu (kJ)",  # H  = D / C  (efficiency metric)
    ]
    ncols = len(col_headers)
    for c, h in enumerate(col_headers, 1):
        ws.cell(row=hdr, column=c, value=h)
    style_header(ws, hdr, ncols)

    # Load corrected data at 1800K
    _, corr_rows = load_csv(CSV_FILES["corrected"])
    rows_1800 = [r for r in corr_rows if abs(float(r["T_K"]) - 1800) < 1]

    # Sort by corrected dG
    rows_1800.sort(key=lambda r: float(r["dG_corrected_pO2_1atm_kJ"]))

    data_start = hdr + 1
    for i, r in enumerate(rows_1800):
        row = data_start + i
        product = r["product"]
        n_cu = int(r["n_Cu"])
        dG_pure = float(r["dG_pure_kJ"])
        parent = r.get("parent_oxide", "")

        # A: Product name (text)
        ws.cell(row=row, column=1, value=product)
        # B: Parent oxide (text)
        ws.cell(row=row, column=2, value=parent)
        # C: n_Cu (integer, raw)
        ws.cell(row=row, column=3, value=n_cu)
        # D: dG_pure (raw TC-Python number)
        ws.cell(row=row, column=4, value=dG_pure)
        fmt_number(ws, row, 4, "+0.0;-0.0;0.0")
        # E: Cu penalty = n_Cu * penalty_per_atom  FORMULA
        #    References $C$8 (the penalty per atom from parameter section)
        ws.cell(row=row, column=5, value=f"=C{row}*$C$8")
        fmt_number(ws, row, 5, "+0.0;-0.0;0.0")
        # F: dG_eff = dG_pure + penalty  FORMULA
        ws.cell(row=row, column=6, value=f"=D{row}+E{row}")
        fmt_number(ws, row, 6, "+0.0;-0.0;0.0")
        # G: Verdict  FORMULA
        ws.cell(row=row, column=7,
                value=f'=IF(F{row}<-10,"ROBUST",'
                      f'IF(F{row}<0,"MARGINAL",'
                      f'IF(F{row}<15,"UNCERTAIN","UNFAVORABLE")))')
        # H: dG per Cu atom = dG_pure / n_Cu  FORMULA
        ws.cell(row=row, column=8, value=f"=D{row}/C{row}")
        fmt_number(ws, row, 8, "+0.0;-0.0;0.0")

    last_data = data_start + len(rows_1800) - 1
    alt_shading(ws, data_start, last_data, ncols)
    center_all(ws, hdr, last_data, ncols)
    apply_borders(ws, hdr, last_data, ncols)

    # ── Section 3: Sensitivity — gamma_Cu sweep ──
    SENS_START = last_data + 2
    ws.cell(row=SENS_START, column=1,
            value="\u03b3_Cu SENSITIVITY ANALYSIS").font = \
        Font(bold=True, size=12)
    ws.merge_cells(f"A{SENS_START}:H{SENS_START}")

    ws.cell(row=SENS_START + 1, column=1,
            value="Change \u03b3_Cu in cell C4 above to see all corrections "
                  "update automatically.").font = note_font
    ws.merge_cells(f"A{SENS_START+1}:H{SENS_START+1}")

    sens_hdr = SENS_START + 2
    sens_cols = [
        "\u03b3_Cu",       # A
        "a_Cu",            # B  = gamma * X_Cu
        "Penalty/Cu (kJ)", # C  = -R*T*LN(a_Cu)/1000
    ]
    # Add one column per product for dG_eff
    product_names = [r["product"] for r in rows_1800]
    for p in product_names:
        sens_cols.append(f"\u0394G_eff {p}")
    sens_ncols = len(sens_cols)

    for c, h in enumerate(sens_cols, 1):
        ws.cell(row=sens_hdr, column=c, value=h)
    style_header(ws, sens_hdr, sens_ncols)

    # Gamma values to sweep
    gamma_values = [3, 4, 5, 6, 7, 8, 8.5, 9, 10, 11, 12, 13, 15]

    for j, gv in enumerate(gamma_values):
        row = sens_hdr + 1 + j
        # A: gamma value (raw)
        ws.cell(row=row, column=1, value=gv)
        fmt_number(ws, row, 1, "0.0")
        # B: a_Cu = gamma * X_Cu  FORMULA
        ws.cell(row=row, column=2, value=f"=A{row}*$C$3")
        fmt_number(ws, row, 2, "0.0000")
        # C: penalty per Cu = -R*T*LN(a_Cu)/1000  FORMULA
        ws.cell(row=row, column=3, value=f"=-$C$7*$C$6*LN(B{row})/1000")
        fmt_number(ws, row, 3, "+0.0;-0.0")

        # D onward: dG_eff for each product = dG_pure + n_Cu * penalty
        for k, r in enumerate(rows_1800):
            col = 4 + k
            n_cu = int(r["n_Cu"])
            dG_pure = float(r["dG_pure_kJ"])
            # dG_eff = dG_pure + n_Cu * penalty  FORMULA
            # dG_pure is a constant, n_Cu is a constant, penalty is in col C
            ws.cell(row=row, column=col,
                    value=f"={dG_pure}+{n_cu}*C{row}")
            fmt_number(ws, row, col, "+0.0;-0.0;0.0")

    sens_last = sens_hdr + len(gamma_values)
    alt_shading(ws, sens_hdr + 1, sens_last, sens_ncols)
    center_all(ws, sens_hdr, sens_last, sens_ncols)
    apply_borders(ws, sens_hdr, sens_last, sens_ncols)

    # Conditional formatting: highlight negative dG_eff cells in green
    from openpyxl.formatting.rule import CellIsRule
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE",
                             fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE",
                           fill_type="solid")
    for k in range(len(rows_1800)):
        col_letter = get_column_letter(4 + k)
        cell_range = f"{col_letter}{sens_hdr+1}:{col_letter}{sens_last}"
        ws.conditional_formatting.add(
            cell_range,
            CellIsRule(operator="lessThan", formula=["0"],
                       fill=green_fill)
        )
        ws.conditional_formatting.add(
            cell_range,
            CellIsRule(operator="greaterThanOrEqual", formula=["0"],
                       fill=red_fill)
        )

    # Also conditionally format the main table column F
    main_range = f"F{data_start}:F{last_data}"
    ws.conditional_formatting.add(
        main_range,
        CellIsRule(operator="lessThan", formula=["0"], fill=green_fill)
    )
    ws.conditional_formatting.add(
        main_range,
        CellIsRule(operator="greaterThanOrEqual", formula=["0"],
                   fill=red_fill)
    )

    # Column widths
    auto_width(ws, sens_ncols)
    # Override some columns for readability
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["D"].width = 12

    # V2O5 note at bottom
    note_row = sens_last + 2
    ws.cell(row=note_row, column=1,
            value="NOTE: V\u2082O\u2085 can form either Cu\u2083V\u2082O\u2088 "
                  "(3 Cu, \u0394G\u00b0 = -109.2 kJ) or CuV\u2082O\u2086 "
                  "(1 Cu, \u0394G\u00b0 = -45.0 kJ). Under dilute Cu "
                  "conditions, CuV\u2082O\u2086 is preferred because the "
                  "3\u00d7 penalty makes Cu\u2083V\u2082O\u2088 unfavorable."
            ).font = note_font
    ws.merge_cells(f"A{note_row}:H{note_row}")

    note_row2 = note_row + 1
    ws.cell(row=note_row2, column=1,
            value="Bureau of Mines (1960s) used ~1 wt% Cu (a_Cu \u2248 0.13), "
                  "where both V\u2082O\u2085 products are favorable. Their "
                  "40-60% removal is consistent with our model."
            ).font = note_font
    ws.merge_cells(f"A{note_row2}:H{note_row2}")

    print(f"  Sheet 3: Activity_Correction ({len(rows_1800)} products, "
          f"{len(gamma_values)} gamma points, ALL formulas)")
    return ws


# =====================================================================
# Sheet 4: Candidate_Ranking
# =====================================================================
def build_candidate_ranking(wb):
    ws = wb.create_sheet("Candidate_Ranking")

    # We need the dG_vs_T sheet reference to pull 1800K values.
    # First figure out which row in Sheet 1 has T_K = 1800.
    # Sheet 1 temps are sorted ascending starting at row 2.
    # Load temps to find the row index.
    _, dg_rows = load_csv(CSV_FILES["dG_top6"])
    temps_sorted = sorted(set(int(r["T_K"]) for r in dg_rows))
    try:
        row_1800 = temps_sorted.index(1800) + 2  # +2 for 1-indexed + header
    except ValueError:
        row_1800 = None

    # Also find products in column order (same order as build_dG_vs_T)
    products_ordered = []
    seen = set()
    for r in dg_rows:
        p = r["product"]
        if p not in seen:
            products_ordered.append(p)
            seen.add(p)
    # Map product -> column letter in dG_vs_T (C, D, E, ...)
    prod_col = {}
    for i, p in enumerate(products_ordered):
        prod_col[p] = get_column_letter(3 + i)

    # Candidate data (from CLAUDE.md ternary results and dG_vs_T data)
    candidates = [
        ("CuFe2O4",  "SPINEL#1",       1700, "Strong"),
        ("Cu3V2O8",  "(liquid)",        None, "Strong"),
        ("CuMn2O4",  "SPINEL#1",       1700, "Strong"),
        ("Cu2SiO4",  "(liquid)",        None, "Strong"),
        ("CuB2O4",   "CUB2O4#1",       1300, "Strong"),
        ("CuAl2O4",  "SPINEL#1",       1550, "Moderate"),
    ]

    headers = [
        "Product",
        "dG at 1800 K (kJ)",
        "Named Phase",
        "Stable To (K)",
        "Tier",
    ]
    ncols = len(headers)
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header(ws, 1, ncols)

    for i, (prod, phase, stable_to, tier) in enumerate(candidates, start=2):
        ws.cell(row=i, column=1, value=prod)

        # Column B: dG at 1800K — cross-sheet formula referencing dG_vs_T
        if row_1800 is not None and prod in prod_col:
            col_l = prod_col[prod]
            ws.cell(
                row=i, column=2,
                value=f"=dG_vs_T!{col_l}{row_1800}"
            )
        else:
            ws.cell(row=i, column=2, value="N/A")
        fmt_number(ws, i, 2, "+0.00;-0.00;0.00")

        ws.cell(row=i, column=3, value=phase)
        ws.cell(row=i, column=4, value=stable_to if stable_to else "N/A (liquid)")
        ws.cell(row=i, column=5, value=tier)

    last_row = 1 + len(candidates)
    alt_shading(ws, 2, last_row, ncols)
    center_all(ws, 1, last_row, ncols)
    apply_borders(ws, 1, last_row, ncols)
    auto_width(ws, ncols)

    print(f"  Sheet 3: Candidate_Ranking ({len(candidates)} candidates, "
          f"cross-sheet dG formulas)")
    return ws


# =====================================================================
# Sheet 4: Slag_Effects
# =====================================================================
def build_slag_effects(wb):
    ws = wb.create_sheet("Slag_Effects")

    _, rows = load_csv(CSV_FILES["slag"])
    if not rows:
        ws.cell(row=1, column=1, value="Data not available")
        return ws

    # Columns: system, ratio_label, ratio_value, a_Cu, stable_phases
    # Add formula column: % change from baseline (first row of each system)
    headers = [
        "System",
        "Ratio Label",
        "Ratio Value",
        "a_Cu",
        "Stable Phases",
        "% Change from Baseline",
    ]
    ncols = len(headers)
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header(ws, 1, ncols)

    # Track the first row of each system for the baseline reference
    system_baseline_row = {}  # system -> Excel row number of first entry
    current_system = None

    for i, r in enumerate(rows, start=2):
        sys = r["system"]
        ws.cell(row=i, column=1, value=sys)
        ws.cell(row=i, column=2, value=r["ratio_label"])
        ws.cell(row=i, column=3, value=_try_float(r["ratio_value"]))
        fmt_number(ws, i, 3, "0.0")
        ws.cell(row=i, column=4, value=_try_float(r["a_Cu"]))
        fmt_number(ws, i, 4, "0.0000E+00")
        ws.cell(row=i, column=5, value=r["stable_phases"])

        # % change formula: =(D{i} - D{baseline}) / D{baseline} * 100
        if sys not in system_baseline_row:
            system_baseline_row[sys] = i
            # First row of this system: % change = 0 (it IS the baseline)
            ws.cell(row=i, column=6, value=0.0)
        else:
            base = system_baseline_row[sys]
            ws.cell(row=i, column=6,
                    value=f"=(D{i}-D${base})/D${base}*100")
        fmt_number(ws, i, 6, "+0.00;-0.00;0.00")

    last_row = 1 + len(rows)

    # Group shading by system
    # Find group boundaries
    grp_idx = 0
    prev_sys = None
    for r in range(2, last_row + 1):
        sys_val = ws.cell(row=r, column=1).value
        if sys_val != prev_sys:
            grp_idx += 1
            prev_sys = sys_val
        fill = WHITE_FILL if grp_idx % 2 == 1 else GRAY_FILL
        apply_row_fill(ws, r, ncols, fill)

    center_all(ws, 1, last_row, ncols)
    apply_borders(ws, 1, last_row, ncols)
    auto_width(ws, ncols)

    print(f"  Sheet 4: Slag_Effects ({len(rows)} rows, "
          f"formula col F = % change from baseline)")
    return ws


# =====================================================================
# Sheet 5: Cu_Activity
# =====================================================================
def build_cu_activity(wb):
    ws = wb.create_sheet("Cu_Activity")

    _, rows = load_csv(CSV_FILES["activity"])
    if not rows:
        ws.cell(row=1, column=1, value="Data not available")
        return ws

    headers = ["System", "Oxide", "T (K)", "X_Cu", "X_M", "X_O",
               "a_Cu", "Stable Phases"]
    ncols = len(headers)
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header(ws, 1, ncols)

    for i, r in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=r["system"])
        ws.cell(row=i, column=2, value=r["oxide"])
        ws.cell(row=i, column=3, value=_try_float(r["T_K"]))
        ws.cell(row=i, column=4, value=_try_float(r["X_Cu"]))
        fmt_number(ws, i, 4, "0.000000")
        ws.cell(row=i, column=5, value=_try_float(r["X_M"]))
        fmt_number(ws, i, 5, "0.000000")
        ws.cell(row=i, column=6, value=_try_float(r["X_O"]))
        fmt_number(ws, i, 6, "0.000000")
        ws.cell(row=i, column=7, value=_try_float(r["a_Cu"]))
        fmt_number(ws, i, 7, "0.0000E+00")
        ws.cell(row=i, column=8, value=r["stable_phases"])

    last_row = 1 + len(rows)

    # Group shading by system
    grp_idx = 0
    prev_sys = None
    for r in range(2, last_row + 1):
        sys_val = ws.cell(row=r, column=1).value
        if sys_val != prev_sys:
            grp_idx += 1
            prev_sys = sys_val
        fill = WHITE_FILL if grp_idx % 2 == 1 else GRAY_FILL
        apply_row_fill(ws, r, ncols, fill)

    center_all(ws, 1, last_row, ncols)
    apply_borders(ws, 1, last_row, ncols)
    auto_width(ws, ncols)

    print(f"  Sheet 5: Cu_Activity ({len(rows)} rows)")
    return ws


# =====================================================================
# Sheet 6: Data_Sources
# =====================================================================
def build_data_sources(wb):
    ws = wb.create_sheet("Data_Sources")

    headers = ["CSV File", "Generated By", "Database", "Description",
               "Date Extracted"]
    ncols = len(headers)
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header(ws, 1, ncols)

    sources = [
        [
            "dG_vs_T_top6.csv",
            "simulations/tcpython/dG_vs_T_top6.py",
            "TCOX14",
            "Fine-resolution dG vs T for top 6 ternary candidates "
            "(25 K steps, 800-1900 K)",
            "2026-03-13",
        ],
        [
            "cufe2o4_alternative_reaction.csv",
            "simulations/tcpython/cufe2o4_alternative_reaction.py",
            "TCOX14",
            "CuFe2O4 formation decomposed into Cu-capture vs "
            "Fe-oxidation components (50 K steps, 800-1900 K)",
            "2026-03-13",
        ],
        [
            "cufe2o4_decomposition_results.csv",
            "screening/analyze_cufe2o4_decomposition.py",
            "(derived from cufe2o4_alternative_reaction.csv)",
            "Pre-computed Fe-oxidation and Cu-capture percentages",
            "2026-03-13",
        ],
        [
            "slag_composition_effects.csv",
            "simulations/tcpython/slag_composition_effects.py",
            "TCOX14",
            "Cu activity in quaternary slag systems "
            "(Cu-Mn-Si-O, Cu-Al-Si-O) vs basicity ratio at 1800 K",
            "2026-03-13",
        ],
        [
            "cu_activity_vs_oxide.csv",
            "simulations/tcpython/cu_activity_vs_oxide.py",
            "TCOX14",
            "Cu activity sweep at 1800 K across 4 ternary systems "
            "(Cu-Al-O, Cu-Mn-O, Cu-Fe-O, Cu-V-O), 20 compositions each",
            "2026-03-13",
        ],
    ]

    for i, src in enumerate(sources, start=2):
        for c, val in enumerate(src, 1):
            ws.cell(row=i, column=c, value=val)

    last_row = 1 + len(sources)
    alt_shading(ws, 2, last_row, ncols)
    center_all(ws, 1, last_row, ncols)
    apply_borders(ws, 1, last_row, ncols)
    auto_width(ws, ncols)

    print(f"  Sheet 6: Data_Sources ({len(sources)} entries)")
    return ws


# =====================================================================
# main
# =====================================================================
def main():
    print("=" * 70)
    print("Building Validation_Results.xlsx  (traceable cell formulas)")
    print("=" * 70)

    wb = openpyxl.Workbook()

    build_dG_vs_T(wb)
    build_cufe2o4_decomp(wb)
    build_activity_correction(wb)
    build_candidate_ranking(wb)
    build_slag_effects(wb)
    build_cu_activity(wb)
    build_data_sources(wb)

    wb.save(OUTPUT_FILE)
    print(f"\nSaved: {OUTPUT_FILE}")
    print("=" * 70)


if __name__ == "__main__":
    main()
